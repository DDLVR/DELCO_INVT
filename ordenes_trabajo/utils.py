"""
Utilidades para órdenes de trabajo: importación masiva, exportación,
detección de duplicados e integración con informes de clientes.
"""
import json
import os
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

import openpyxl
from django.conf import settings
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Count, Q
from django.utils import timezone

from clientes.models import Cliente
from importaciones.models import ImportacionExcel, ImportacionExcelError
from usuarios.models import Usuario

from .models import InformeCliente, OrdenTrabajo
from .services import validate_ot_for_creation

DIAS_ALERTA_DUPLICADO = 14

TIPO_TRABAJO_MAP = {
    'INSTALACION': 'INSTALACION',
    'INSTALACIÓN': 'INSTALACION',
    'INSTALACION MEDIDOR': 'INSTALACION',
    'CAMBIO': 'CAMBIO',
    'CAMBIO DE EQUIPO': 'CAMBIO',
    'RETIRO': 'RETIRO',
    'MANTENCION': 'MANTENCION',
    'MANTENCIÓN': 'MANTENCION',
    'MANTENIMIENTO': 'MANTENCION',
    'REPARACION': 'REPARACION',
    'REPARACIÓN': 'REPARACION',
    'INSPECCION': 'INSPECCION',
    'INSPECCIÓN': 'INSPECCION',
    'CONFIGURACION': 'CONFIGURACION',
    'CONFIGURACIÓN': 'CONFIGURACION',
    'OTRO': 'OTRO',
}

ESTADO_MAP = {
    codigo: codigo for codigo, _ in OrdenTrabajo.ESTADO_CHOICES
}
for codigo, etiqueta in OrdenTrabajo.ESTADO_CHOICES:
    ESTADO_MAP[etiqueta.upper()] = codigo
    clave = unicodedata.normalize('NFD', etiqueta.upper())
    clave = ''.join(c for c in clave if unicodedata.category(c) != 'Mn')
    ESTADO_MAP[clave] = codigo

# Plantilla asignación de trabajo (técnicos) + columnas extra (proyecto y fechas).
# Los alias antiguos se conservan para no romper Excel exportados antes.
COLUMNAS_ORDEN = {
    'numero_cliente': (
        'numero_cliente', 'numero cliente', 'n cliente', 'n° cliente',
        'nº cliente', 'id cliente', 'nro cliente', 'nro. cliente',
        'cliente',
    ),
    'solicitud': ('solicitud', 'n solicitud', 'nro solicitud', 'numero solicitud', 'nº solicitud'),
    'titulo': ('titulo', 'título', 'titulo trabajo', 'asunto'),
    'descripcion': ('descripcion', 'descripción', 'detalle'),
    'tipo_trabajo': (
        'tipo_trabajo', 'tipo trabajo', 'tipo', 'actividad', 'trabajo',
    ),
    'tecnico': (
        'tecnico', 'técnico', 'responsable', 'tecnico responsable',
        'tecnico_responsable', 'técnico responsable', 'asignado a',
    ),
    'estado': ('estado',),
    'observaciones_tecnicas': ('observaciones tecnicas', 'observaciones_tecnicas'),
    'proyecto_carga_administrativa': (
        'proyecto', 'proyecto carga', 'proyecto / carga administrativa',
        'proyecto carga administrativa', 'carga administrativa', 'proyecto_carga',
        'proyecto_carga_administrativa',
    ),
    'id_orden': ('id orden', 'id_orden'),
    'direccion_cliente': ('direccion cliente', 'direccion_cliente', 'direccion'),
    'comuna': ('comuna',),
    'nombre_cliente': ('nombre', 'nombre cliente', 'customer name', 'customer_name'),
    'medidor': ('medidor', 'serie medidor', 'medidor serie'),
    'marca': ('marca', 'marca medidor'),
    'ip': ('ip',),
    'puerto': ('puerto', 'port'),
    'modem': ('modem', 'serie modem', 'modem serie'),
    'fecha_trabajo': (
        'fecha', 'fecha trabajo', 'fecha asignacion', 'fecha asignación',
        'fecha_asignacion',
    ),
}


def _normalizar_texto(valor) -> str:
    if valor is None:
        return ''
    if isinstance(valor, bool):
        return 'Si' if valor else 'No'
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    texto = str(valor).strip()
    if texto.endswith('.0'):
        base = texto[:-2]
        if base.isdigit():
            return base
    return texto


def _limpiar_header(texto) -> str:
    clave = _normalizar_texto(texto).lower()
    if clave.startswith('\ufeff'):
        clave = clave.lstrip('\ufeff')
    clave = unicodedata.normalize('NFD', clave)
    return ''.join(c for c in clave if unicodedata.category(c) != 'Mn')


def _variantes_header(clave: str):
    variantes = {clave}
    variantes.add(clave.replace(' ', '_'))
    variantes.add(clave.replace('_', ' '))
    return variantes


def _mapear_headers(headers) -> Dict[str, int]:
    indice = {}
    for i, header in enumerate(headers):
        if header is None:
            continue
        clave_base = _limpiar_header(header)
        for campo, alias in COLUMNAS_ORDEN.items():
            if campo in indice:
                continue
            for variante in _variantes_header(clave_base):
                if variante in alias:
                    indice[campo] = i
                    break
    return indice


def _valor_fila(valores, indice, campo, default=''):
    pos = indice.get(campo)
    if pos is None or pos >= len(valores):
        return default
    return _normalizar_texto(valores[pos])


def _valor_fila_raw(valores, indice, campo, default=None):
    pos = indice.get(campo)
    if pos is None or pos >= len(valores):
        return default
    return valores[pos]


def _resolver_tipo_trabajo(texto: str) -> str:
    clave = _normalizar_texto(texto).upper()
    if not clave:
        return 'INSTALACION'
    clave_sin_acentos = unicodedata.normalize('NFD', clave)
    clave_sin_acentos = ''.join(c for c in clave_sin_acentos if unicodedata.category(c) != 'Mn')
    if clave in dict(OrdenTrabajo.TIPO_TRABAJO_CHOICES):
        return clave
    if clave_sin_acentos in dict(OrdenTrabajo.TIPO_TRABAJO_CHOICES):
        return clave_sin_acentos
    return TIPO_TRABAJO_MAP.get(clave, TIPO_TRABAJO_MAP.get(clave_sin_acentos, 'OTRO'))


def _resolver_estado(texto: str) -> Optional[str]:
    clave = _normalizar_texto(texto).upper()
    if not clave:
        return None
    clave_sin_acentos = unicodedata.normalize('NFD', clave)
    clave_sin_acentos = ''.join(c for c in clave_sin_acentos if unicodedata.category(c) != 'Mn')
    return ESTADO_MAP.get(clave) or ESTADO_MAP.get(clave_sin_acentos)


def _parse_id_orden(valor) -> Optional[int]:
    texto = _normalizar_texto(valor)
    if not texto:
        return None
    try:
        return int(float(texto))
    except (ValueError, TypeError):
        return None


def _parse_fecha_excel(valor) -> Optional[datetime]:
    """Convierte celda Excel (datetime, date o texto dd/mm/aaaa) a datetime aware."""
    if valor is None or valor == '':
        return None
    dt = None
    if isinstance(valor, datetime):
        dt = valor
    elif isinstance(valor, date):
        dt = datetime.combine(valor, time.min)
    else:
        texto = _normalizar_texto(valor)
        if not texto:
            return None
        for fmt in ('%d/%m/%Y %H:%M:%S', '%d/%m/%Y %H:%M', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
            try:
                dt = datetime.strptime(texto, fmt)
                break
            except ValueError:
                continue
    if dt is None:
        return None
    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt, timezone.get_current_timezone())
    return dt


def _buscar_medidor_por_serie(serie: str):
    from inventario.models import Medidor

    clave = _normalizar_texto(serie)
    if not clave:
        return None
    return Medidor.objects.filter(serie__iexact=clave, eliminado=False).first()


def _buscar_modem_por_serie(serie: str):
    from inventario.models import Modem

    clave = _normalizar_texto(serie)
    if not clave:
        return None
    return Modem.objects.filter(serie__iexact=clave, eliminado=False).first()


def _rellenar_cliente_si_vacio(cliente: Cliente, campos: Dict[str, str]) -> None:
    """Completa datos del cliente solo si el campo está vacío (no pisa lo existente)."""
    update = []
    mapping = (
        ('customer_name', 'customer_name', 255),
        ('direccion', 'direccion', 255),
        ('installation_address', 'installation_address', 255),
        ('comuna', 'comuna', 100),
        ('ip', 'ip', 45),
        ('puerto', 'puerto', 50),
        ('modem', 'modem', 255),
        ('meter_serial_n_1', 'meter_serial_n_1', 100),
        ('meter_manufacturer_id', 'meter_manufacturer_id', 255),
    )
    for attr, key, maxlen in mapping:
        nuevo = (campos.get(key) or '').strip()
        if not nuevo:
            continue
        actual = getattr(cliente, attr, None)
        if actual is None or str(actual).strip() == '':
            setattr(cliente, attr, nuevo[:maxlen])
            update.append(attr)
    if update:
        cliente.save(update_fields=update)


def _componer_descripcion_asignacion(
    descripcion: str,
    solicitud: str,
    nombre: str,
    direccion: str,
    comuna: str,
    medidor: str,
    marca: str,
    ip: str,
    puerto: str,
    modem: str,
) -> str:
    if (descripcion or '').strip():
        return descripcion
    lineas = []
    if solicitud:
        lineas.append(f'Solicitud: {solicitud}')
    if nombre:
        lineas.append(f'Nombre: {nombre}')
    if direccion:
        lineas.append(f'Dirección: {direccion}')
    if comuna:
        lineas.append(f'Comuna: {comuna}')
    if medidor:
        lineas.append(f'Medidor: {medidor}')
    if marca:
        lineas.append(f'Marca: {marca}')
    if ip or puerto:
        lineas.append('IP / Puerto: {} / {}'.format(ip or '—', puerto or '—'))
    if modem:
        lineas.append(f'Módem: {modem}')
    return '\n'.join(lineas)


def _extraer_correlativo_moreapp(observaciones: str) -> Optional[str]:
    match = re.search(r'correlativo:\s*(\d+)', observaciones or '', flags=re.I)
    return match.group(1) if match else None


def _buscar_orden_existente_en_importacion(
    cliente: Cliente,
    titulo: str,
    observaciones: str,
    orden_id: Optional[int] = None,
) -> Optional[OrdenTrabajo]:
    """Evita crear OT duplicadas al reimportar el mismo Excel."""
    if orden_id:
        encontrada = OrdenTrabajo.objects.filter(pk=orden_id, eliminado=False).first()
        if encontrada:
            return encontrada

    titulo_norm = (titulo or '')[:200]
    if titulo_norm:
        encontrada = OrdenTrabajo.objects.filter(
            cliente=cliente,
            titulo=titulo_norm,
            eliminado=False,
        ).order_by('-id').first()
        if encontrada:
            return encontrada

    correlativo = _extraer_correlativo_moreapp(observaciones)
    if correlativo:
        encontrada = OrdenTrabajo.objects.filter(
            cliente=cliente,
            eliminado=False,
            observaciones_tecnicas__icontains=f'correlativo: {correlativo}',
        ).order_by('-id').first()
        if encontrada:
            return encontrada

    return None


def _resolver_tecnico(texto: str) -> Optional[Usuario]:
    """Busca técnico por nombre_interno, nombre, email o RUT. Nunca lanza error."""
    nombre = _normalizar_texto(texto)
    if not nombre:
        return None
    qs = Usuario.objects.filter(rol='TECNICO', is_active=True)
    for lookup in (
        {'nombre_interno__iexact': nombre},
        {'nombre__iexact': nombre},
        {'email__iexact': nombre},
        {'rut__iexact': nombre},
    ):
        tecnico = qs.filter(**lookup).first()
        if tecnico:
            return tecnico
    return None


def _obtener_cliente_existente(numero_cliente: str) -> Cliente:
    """
    Busca un cliente existente. NO crea clientes nuevos.
    La carga masiva de OT exige que el cliente ya exista en BD.
    """
    numero = (str(numero_cliente).strip() if numero_cliente is not None else '')
    if not numero:
        raise ValueError('Numero Cliente es obligatorio')

    cliente = Cliente.objects.filter(numero_cliente=numero, activo=True).first()
    if not cliente:
        cliente = Cliente.objects.filter(numero_cliente__iexact=numero, activo=True).first()
    if not cliente:
        raise ValueError(
            f'El cliente «{numero}» no existe en la base de datos. '
            'La orden no se importó: cree el cliente antes o corrija el número.'
        )
    return cliente


def _obtener_o_crear_cliente(numero_cliente: str, valores=None, indice=None) -> Cliente:
    """Compatibilidad: ya no crea clientes; delega en _obtener_cliente_existente."""
    return _obtener_cliente_existente(numero_cliente)


def _aplicar_tecnico_a_orden(orden: OrdenTrabajo, tecnico: Optional[Usuario], fecha_asignacion=None) -> None:
    if tecnico:
        orden.tecnico_responsable = tecnico
        if orden.estado == 'CREADA':
            orden.estado = 'ASIGNADA'
        if fecha_asignacion and not orden.fecha_asignacion:
            orden.fecha_asignacion = fecha_asignacion
        elif not orden.fecha_asignacion:
            orden.fecha_asignacion = timezone.now()


def _fila_a_texto(headers, valores):
    data = {}
    for i, valor in enumerate(valores):
        nombre = headers[i] if i < len(headers) else f'Columna_{i + 1}'
        data[str(nombre)] = valor
    return json.dumps(data, ensure_ascii=False, default=str)


def detectar_duplicado_orden(
    cliente,
    exclude_orden_id=None,
    dias: int = DIAS_ALERTA_DUPLICADO,
) -> Tuple[bool, str]:
    """
    Alerta si existe otra orden activa/reciente para el mismo cliente
    dentro de la ventana configurada (14 días por defecto).
    """
    if not cliente:
        return False, ''

    desde = timezone.now() - timedelta(days=dias)
    estados_relevantes = list(OrdenTrabajo.ESTADOS_ABIERTOS) + [
        'REALIZADA',
        'REALIZADA_PENDIENTE_COMPROBACION',
        'PENDIENTE_VALIDACION',
    ]

    qs = OrdenTrabajo.objects.filter(
        cliente=cliente,
        fecha_creacion__gte=desde,
        estado__in=estados_relevantes,
        eliminado=False,
    ).exclude(estado='CANCELADA')

    if exclude_orden_id:
        qs = qs.exclude(pk=exclude_orden_id)

    anterior = qs.order_by('-fecha_creacion').first()
    if not anterior:
        return False, ''

    dias_diff = (timezone.now() - anterior.fecha_creacion).days
    desc = (
        f'Posible trabajo duplicado — Cliente: {cliente.numero_cliente} | '
        f'Orden anterior: #{anterior.id} ({anterior.get_estado_display()}) | '
        f'Creada: {anterior.fecha_creacion.strftime("%d/%m/%Y %H:%M")} | '
        f'Hace {dias_diff} día(s) (ventana: {dias} días)'
    )
    return True, desc


def aplicar_alerta_duplicado(orden: OrdenTrabajo) -> None:
    if not orden.cliente_id:
        return
    tiene_alerta, desc = detectar_duplicado_orden(orden.cliente, exclude_orden_id=orden.pk)
    if tiene_alerta:
        orden.alerta_duplicado = True
        orden.descripcion_alerta_duplicado = desc
        orden.save(update_fields=['alerta_duplicado', 'descripcion_alerta_duplicado'])


def importar_ordenes_excel(archivo, usuario) -> ImportacionExcel:
    importacion = ImportacionExcel.objects.create(
        tipo='ORDENES_TRABAJO',
        archivo_original=getattr(archivo, 'name', 'Upload'),
        usuario=usuario,
    )

    try:
        nombre_archivo = getattr(archivo, 'name', '') or ''
        if nombre_archivo.lower().endswith('.xls') and not nombre_archivo.lower().endswith('.xlsx'):
            raise ValueError(
                'Formato .xls no soportado. Guarda el archivo como Excel (.xlsx) e intenta de nuevo.'
            )

        if hasattr(archivo, 'seek'):
            archivo.seek(0)
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        indice = _mapear_headers(headers)

        if 'numero_cliente' not in indice:
            raise ValueError(
                'El Excel debe incluir la columna "CLIENTE" (Nº cliente). '
                f'Columnas detectadas: {", ".join(str(h) for h in headers if h)}'
            )

        # ── Validación previa: leer todas las filas y verificar clientes ──
        filas_trabajo = []
        errores_previos = []
        clientes_vistos = {}
        clientes_ok = set()
        clientes_faltantes = set()

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            valores = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != '' for v in valores):
                continue

            numero_cliente_raw = _valor_fila(valores, indice, 'numero_cliente')
            if numero_cliente_raw and _limpiar_header(numero_cliente_raw) in COLUMNAS_ORDEN['numero_cliente']:
                continue

            numero = (numero_cliente_raw or '').strip() if numero_cliente_raw is not None else ''
            if not numero:
                errores_previos.append({
                    'fila': idx,
                    'motivo': 'Numero Cliente es obligatorio',
                    'data': _fila_a_texto(headers, valores),
                    'numero': '',
                })
                continue

            if numero not in clientes_vistos:
                try:
                    clientes_vistos[numero] = _obtener_cliente_existente(numero)
                    clientes_ok.add(numero)
                except ValueError as exc:
                    clientes_vistos[numero] = None
                    clientes_faltantes.add(numero)
                    errores_previos.append({
                        'fila': idx,
                        'motivo': str(exc),
                        'data': _fila_a_texto(headers, valores),
                        'numero': numero,
                    })
                    continue
            elif clientes_vistos[numero] is None:
                clientes_faltantes.add(numero)
                errores_previos.append({
                    'fila': idx,
                    'motivo': (
                        f'El cliente «{numero}» no existe en la base de datos. '
                        'La orden no se importó: cree el cliente antes o corrija el número.'
                    ),
                    'data': _fila_a_texto(headers, valores),
                    'numero': numero,
                })
                continue

            filas_trabajo.append((idx, valores, numero))

        if clientes_faltantes or any(e['motivo'].startswith('Numero Cliente') for e in errores_previos):
            # No insertar ninguna OT si hay clientes inexistentes u obligatoriedad fallida
            for err in errores_previos:
                ImportacionExcelError.objects.create(
                    importacion=importacion,
                    numero_fila=err['fila'],
                    motivo=err['motivo'],
                    data_cruda=err['data'],
                )
            importacion.total_filas = len(filas_trabajo) + len(errores_previos)
            importacion.exitosas = 0
            importacion.fallidas = len(errores_previos)
            importacion.estado = 'ERROR'
            lista_faltantes = ', '.join(sorted(clientes_faltantes)[:20])
            extra = f' (+{len(clientes_faltantes) - 20} más)' if len(clientes_faltantes) > 20 else ''
            importacion.observaciones = (
                f'Validación previa fallida. No se importó ninguna orden. '
                f'Total filas revisadas: {importacion.total_filas}. '
                f'Clientes encontrados: {len(clientes_ok)}. '
                f'Clientes inexistentes: {len(clientes_faltantes)}'
                + (f' ({lista_faltantes}{extra}).' if clientes_faltantes else '.')
                + ' Corrija el Excel o cree los clientes antes de reintentar.'
            )
            importacion.save()
            return importacion

        contador = 0
        exitosas = 0
        fallidas = 0
        alertas = 0
        creadas = 0
        actualizadas = 0

        from clientes.proyecto_historial import asignar_proyecto_al_crear_ot

        for idx, valores, numero_cliente in filas_trabajo:
            contador += 1
            try:
                cliente = clientes_vistos[numero_cliente]

                solicitud = _valor_fila(valores, indice, 'solicitud')
                titulo = (
                    _valor_fila(valores, indice, 'titulo')
                    or solicitud
                    or f'Trabajo — {numero_cliente}'
                )
                nombre_cliente = _valor_fila(valores, indice, 'nombre_cliente')
                direccion = _valor_fila(valores, indice, 'direccion_cliente')
                comuna = _valor_fila(valores, indice, 'comuna')
                serie_medidor = _valor_fila(valores, indice, 'medidor')
                marca = _valor_fila(valores, indice, 'marca')
                ip = _valor_fila(valores, indice, 'ip')
                puerto = _valor_fila(valores, indice, 'puerto')
                serie_modem = _valor_fila(valores, indice, 'modem')
                descripcion = _componer_descripcion_asignacion(
                    _valor_fila(valores, indice, 'descripcion'),
                    solicitud,
                    nombre_cliente,
                    direccion,
                    comuna,
                    serie_medidor,
                    marca,
                    ip,
                    puerto,
                    serie_modem,
                )
                tipo_trabajo = _resolver_tipo_trabajo(_valor_fila(valores, indice, 'tipo_trabajo'))
                tecnico_nombre = _valor_fila(valores, indice, 'tecnico')
                tecnico = _resolver_tecnico(tecnico_nombre)
                estado_import = _resolver_estado(_valor_fila(valores, indice, 'estado'))
                observaciones_tecnicas = _valor_fila(valores, indice, 'observaciones_tecnicas')
                proyecto_carga = _valor_fila(valores, indice, 'proyecto_carga_administrativa')
                orden_id = _parse_id_orden(_valor_fila(valores, indice, 'id_orden'))
                fecha_asignacion_excel = _parse_fecha_excel(
                    _valor_fila_raw(valores, indice, 'fecha_trabajo')
                )
                medidor_obj = _buscar_medidor_por_serie(serie_medidor)
                modem_obj = _buscar_modem_por_serie(serie_modem)

                _rellenar_cliente_si_vacio(cliente, {
                    'customer_name': nombre_cliente,
                    'direccion': direccion,
                    'installation_address': direccion,
                    'comuna': comuna,
                    'ip': ip,
                    'puerto': puerto,
                    'modem': serie_modem,
                    'meter_serial_n_1': serie_medidor,
                    'meter_manufacturer_id': marca,
                })

                with transaction.atomic():
                    orden = None
                    ot_validation = None
                    orden = _buscar_orden_existente_en_importacion(
                        cliente,
                        titulo,
                        observaciones_tecnicas,
                        orden_id,
                    )

                    if orden:
                        orden.cliente = cliente
                        orden.titulo = titulo[:200]
                        orden.descripcion = descripcion
                        orden.tipo_trabajo = tipo_trabajo
                        if observaciones_tecnicas:
                            orden.observaciones_tecnicas = observaciones_tecnicas
                        if proyecto_carga is not None and str(proyecto_carga).strip() != '':
                            orden.proyecto_carga_administrativa = str(proyecto_carga).strip()[:255]
                        # El estado no se fuerza desde Excel (evita saltarse flujo/validación)
                        if medidor_obj:
                            orden.medidor = medidor_obj
                        if modem_obj:
                            orden.modem = modem_obj
                        if tecnico:
                            _aplicar_tecnico_a_orden(orden, tecnico, fecha_asignacion_excel)
                        elif fecha_asignacion_excel and not orden.fecha_asignacion:
                            orden.fecha_asignacion = fecha_asignacion_excel
                        elif not tecnico_nombre:
                            pass
                        orden.save()
                        actualizadas += 1
                    else:
                        ot_validation = validate_ot_for_creation(cliente, tipo_trabajo)
                        if ot_validation.has_blocking_error:
                            raise ValueError('; '.join(ot_validation.errors))

                        orden = OrdenTrabajo(
                            titulo=titulo[:200],
                            descripcion=descripcion,
                            tipo_trabajo=tipo_trabajo,
                            cliente=cliente,
                            creada_por=usuario,
                            estado='CREADA',
                        )
                        if observaciones_tecnicas:
                            orden.observaciones_tecnicas = observaciones_tecnicas
                        if proyecto_carga:
                            orden.proyecto_carga_administrativa = str(proyecto_carga).strip()[:255]
                        elif getattr(cliente, 'proyecto', None):
                            from web.services.filtros_export import es_sin_proyecto
                            if not es_sin_proyecto(cliente.proyecto):
                                orden.proyecto_carga_administrativa = (cliente.proyecto or '')[:255]
                        if medidor_obj:
                            orden.medidor = medidor_obj
                        if modem_obj:
                            orden.modem = modem_obj
                        if tecnico:
                            _aplicar_tecnico_a_orden(orden, tecnico, fecha_asignacion_excel)
                        elif fecha_asignacion_excel:
                            orden.fecha_asignacion = fecha_asignacion_excel
                        orden.save()
                        creadas += 1

                    # Proyecto va al CLIENTE (no solo a la OT)
                    proyecto_para_cliente = (orden.proyecto_carga_administrativa or '').strip()
                    if proyecto_para_cliente:
                        asignar_proyecto_al_crear_ot(
                            cliente,
                            proyecto_para_cliente,
                            usuario=usuario,
                            motivo=f'Importación OT #{orden.pk}',
                        )

                    aplicar_alerta_duplicado(orden)
                    if orden.alerta_duplicado:
                        alertas += 1
                    if ot_validation and ot_validation.warnings:
                        alertas += len(ot_validation.warnings)

                exitosas += 1
            except Exception as exc:
                fallidas += 1
                ImportacionExcelError.objects.create(
                    importacion=importacion,
                    numero_fila=idx,
                    motivo=str(exc),
                    data_cruda=_fila_a_texto(headers, valores),
                )

        importacion.total_filas = contador
        importacion.exitosas = exitosas
        importacion.fallidas = fallidas
        importacion.estado = 'COMPLETADO' if contador > 0 and exitosas > 0 else ('ERROR' if contador == 0 else 'COMPLETADO')
        if contador == 0:
            importacion.observaciones = 'No se encontraron filas con datos para importar.'
        else:
            importacion.observaciones = (
                f'Se procesaron {exitosas} de {contador} filas '
                f'(creadas: {creadas}, actualizadas: {actualizadas}). '
                f'Alertas duplicidad: {alertas}. Fallidas: {fallidas}. '
                f'Clientes resueltos en validación previa: {len(clientes_ok)}.'
            )
        importacion.save()
    except Exception as exc:
        importacion.estado = 'ERROR'
        importacion.observaciones = f'Error en importación: {exc}'
        importacion.save()

    return importacion


def exportar_ordenes_excel(ordenes):
    """Genera workbook Excel en formato Plantilla Asignación de Trabajo (técnicos)."""
    from importaciones.utils import aplicar_estilo_hoja_exportacion

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Asignacion Trabajo'
    ws.append([
        'SOLICITUD',
        'CLIENTE',
        'MEDIDOR',
        'MARCA',
        'NOMBRE',
        'DIRECCION',
        'COMUNA',
        'TECNICO',
        'TRABAJO',
        'IP',
        'PUERTO',
        'MODEM',
        'FECHA',
        'PROYECTO',
        'Fecha Creacion',
        'Fecha Asignacion',
    ])

    for orden in ordenes:
        cliente = orden.cliente
        medidor = orden.medidor
        modem = orden.modem
        ws.append([
            orden.titulo or '',
            cliente.numero_cliente if cliente else '',
            medidor.serie if medidor else (getattr(cliente, 'meter_serial_n_1', None) or ''),
            (medidor.marca if medidor and getattr(medidor, 'marca', None) else '')
            or (getattr(cliente, 'meter_manufacturer_id', None) or ''),
            (getattr(cliente, 'customer_name', None) or '') if cliente else '',
            (cliente.direccion if cliente else '') or (getattr(cliente, 'installation_address', None) or ''),
            cliente.comuna if cliente else '',
            orden.tecnico_responsable.nombre_interno if orden.tecnico_responsable else '',
            orden.get_tipo_trabajo_display(),
            (getattr(cliente, 'ip', None) or (getattr(modem, 'ip', None) if modem else '')) or '',
            (getattr(cliente, 'puerto', None) or (getattr(modem, 'puerto', None) if modem else '')) or '',
            (modem.serie if modem else '') or (getattr(cliente, 'modem', None) or ''),
            orden.fecha_asignacion.strftime('%d/%m/%Y') if orden.fecha_asignacion else (
                orden.fecha_creacion.strftime('%d/%m/%Y') if orden.fecha_creacion else ''
            ),
            orden.proyecto_carga_administrativa or (getattr(cliente, 'proyecto', None) or ''),
            orden.fecha_creacion.strftime('%d/%m/%Y %H:%M') if orden.fecha_creacion else '',
            orden.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if orden.fecha_asignacion else '',
        ])

    # Filtro en Técnico / Trabajo / Proyecto
    aplicar_estilo_hoja_exportacion(ws, auto_filter=True, filter_from_col=8, filter_to_col=14)
    return wb


def guardar_informe_pdf(
    cliente,
    archivo_origen,
    nombre_archivo: str,
    orden=None,
    usuario=None,
    origen: str = 'MANUAL',
    registro_moreapp=None,
) -> InformeCliente:
    """Guarda un PDF en Registros/Evidencias."""
    Path(settings.EVIDENCIAS_ROOT).mkdir(parents=True, exist_ok=True)

    informe = InformeCliente(
        orden=orden,
        cliente=cliente,
        nombre_archivo=nombre_archivo,
        subido_por=usuario,
        origen=origen,
        registro_moreapp=registro_moreapp,
    )

    if hasattr(archivo_origen, 'read'):
        informe.archivo.save(nombre_archivo, archivo_origen, save=False)
    elif isinstance(archivo_origen, (bytes, bytearray)):
        informe.archivo.save(nombre_archivo, ContentFile(archivo_origen), save=False)
    elif isinstance(archivo_origen, str) and os.path.isfile(archivo_origen):
        with open(archivo_origen, 'rb') as fh:
            informe.archivo.save(nombre_archivo, ContentFile(fh.read()), save=False)
    else:
        raise ValueError('Origen de archivo PDF no válido')

    informe.save()
    return informe


def _buscar_pdfs_en_carpeta(ruta_carpeta: str):
    if not ruta_carpeta or not os.path.isdir(ruta_carpeta):
        return []
    return [
        os.path.join(ruta_carpeta, nombre)
        for nombre in os.listdir(ruta_carpeta)
        if nombre.lower().endswith('.pdf')
    ]


def vincular_informe_cliente_a_orden(
    cliente,
    registro_moreapp=None,
    ruta_carpeta: Optional[str] = None,
    usuario=None,
) -> Optional[OrdenTrabajo]:
    """
    Cuando llega un informe con número de cliente, marca la orden abierta
    como REALIZADA_PENDIENTE_COMPROBACION y guarda PDFs encontrados.
    """
    if not cliente:
        return None

    orden = OrdenTrabajo.objects.filter(
        cliente=cliente,
        eliminado=False,
        estado__in=list(OrdenTrabajo.ESTADOS_ABIERTOS) + ['REALIZADA'],
    ).order_by('-fecha_creacion', '-id').first()

    if orden and orden.estado != 'REALIZADA_PENDIENTE_COMPROBACION':
        orden.estado = 'REALIZADA_PENDIENTE_COMPROBACION'
        if not orden.fecha_fin_ejecucion:
            orden.fecha_fin_ejecucion = timezone.now()
        orden.save(update_fields=['estado', 'fecha_fin_ejecucion'])

    if registro_moreapp:
        registro_moreapp.orden = orden
        registro_moreapp.save(update_fields=['orden'])

    pdfs = _buscar_pdfs_en_carpeta(ruta_carpeta or '')
    for pdf_path in pdfs:
        nombre = os.path.basename(pdf_path)
        if InformeCliente.objects.filter(
            cliente=cliente,
            nombre_archivo=nombre,
            registro_moreapp=registro_moreapp,
        ).exists():
            continue
        guardar_informe_pdf(
            cliente=cliente,
            archivo_origen=pdf_path,
            nombre_archivo=nombre,
            orden=orden,
            usuario=usuario,
            origen='MOREAPP' if registro_moreapp else 'SISTEMA',
            registro_moreapp=registro_moreapp,
        )

    return orden


def asignar_ordenes_masivo(ids, tecnico_id, usuario) -> Dict[str, Any]:
    tecnico = Usuario.objects.get(pk=tecnico_id, rol='TECNICO', is_active=True)
    ordenes = OrdenTrabajo.objects.filter(pk__in=ids, eliminado=False)
    actualizadas = 0
    alertas = 0

    for orden in ordenes:
        orden.tecnico_responsable = tecnico
        orden.estado = 'ASIGNADA'
        orden.fecha_asignacion = timezone.now()
        orden.save(update_fields=['tecnico_responsable', 'estado', 'fecha_asignacion'])
        aplicar_alerta_duplicado(orden)
        if orden.alerta_duplicado:
            alertas += 1
        actualizadas += 1

    return {
        'actualizadas': actualizadas,
        'alertas_duplicado': alertas,
        'tecnico': tecnico.nombre_interno,
    }


COLAS_ORDEN = (
    ('sin_asignar', 'Sin asignar'),
    ('en_campo', 'En campo'),
    ('esperando_moreapp', 'Sin informe MoreApp'),
    ('post_moreapp', 'Post-MoreApp'),
    ('validar', 'Por validar'),
    ('pendientes', 'OT pendientes'),
    ('observadas', 'Observadas'),
)

# Trabajos ya ejecutados / cerrados (vista de consulta)
ESTADOS_TERMINADOS = (
    'REALIZADA',
    'VALIDADA',
    'FINALIZADA',
)

ESTADOS_TERMINADOS_LABELS = (
    ('REALIZADA', 'Realizada'),
    ('VALIDADA', 'Validada'),
    ('FINALIZADA', 'Finalizada'),
)


def aplicar_cola_ordenes(qs, cola: str):
    """Filtros rápidos alineados al flujo Delco → técnico → MoreApp → validación."""
    from reportes.services import ESTADOS_PENDIENTES_OT

    if cola == 'sin_asignar':
        return qs.filter(estado='CREADA', tecnico_responsable__isnull=True)
    if cola == 'en_campo':
        return qs.filter(estado__in=['ASIGNADA', 'EN_EJECUCION'])
    if cola == 'esperando_moreapp':
        return qs.annotate(_n_moreapp=Count('sincronizaciones_moreapp')).filter(
            estado__in=['ASIGNADA', 'EN_EJECUCION'],
            _n_moreapp=0,
        )
    if cola == 'post_moreapp':
        return qs.filter(estado='REALIZADA_PENDIENTE_COMPROBACION')
    if cola == 'validar':
        return qs.filter(estado__in=['PENDIENTE_VALIDACION', 'REALIZADA_PENDIENTE_COMPROBACION'])
    if cola == 'pendientes':
        return qs.filter(estado__in=ESTADOS_PENDIENTES_OT)
    if cola == 'observadas':
        return qs.filter(estado='OBSERVADA')
    return qs


def contadores_colas_ordenes(qs) -> Dict[str, int]:
    """Conteos para pestañas de cola operativa en listado de OT."""
    from reportes.services import ESTADOS_PENDIENTES_OT

    base = qs.annotate(_n_moreapp=Count('sincronizaciones_moreapp', distinct=True))
    return {
        'sin_asignar': base.filter(estado='CREADA', tecnico_responsable__isnull=True).count(),
        'en_campo': base.filter(estado__in=['ASIGNADA', 'EN_EJECUCION']).count(),
        'esperando_moreapp': base.filter(
            estado__in=['ASIGNADA', 'EN_EJECUCION'],
            _n_moreapp=0,
        ).count(),
        'post_moreapp': base.filter(estado='REALIZADA_PENDIENTE_COMPROBACION').count(),
        'validar': base.filter(
            estado__in=['PENDIENTE_VALIDACION', 'REALIZADA_PENDIENTE_COMPROBACION']
        ).count(),
        'pendientes': base.filter(estado__in=ESTADOS_PENDIENTES_OT).count(),
        'observadas': base.filter(estado='OBSERVADA').count(),
    }


def paso_operativo_ot(orden, moreapp_count: int = 0, sync_advertencia: bool = False) -> Dict[str, Any]:
    """Guía contextual del siguiente paso sin alterar el flujo acordado con Delco."""
    estado = orden.estado
    paso = {
        'nivel': 'info',
        'titulo': 'Seguimiento operativo',
        'mensaje': '',
        'accion_url': '',
        'accion_label': '',
    }

    if estado == 'CREADA':
        paso.update(
            nivel='warning',
            titulo='Paso 1 — Asignar técnico',
            mensaje='La orden fue creada por Delco. Asigne un técnico responsable para que pueda ir a terreno.',
            accion_url='',
            accion_label='',
        )
    elif estado == 'ASIGNADA':
        paso.update(
            nivel='info',
            titulo='Paso 2 — Iniciar ejecución',
            mensaje='El técnico debe marcar la orden como en ejecución al llegar al cliente.',
        )
    elif estado == 'EN_EJECUCION' and moreapp_count == 0:
        paso.update(
            nivel='primary',
            titulo='Paso 3 — Esperando informe MoreApp',
            mensaje='El técnico debe completar el formulario en MoreApp al terminar el trabajo en terreno.',
        )
    elif estado == 'EN_EJECUCION' and moreapp_count > 0:
        paso.update(
            nivel='info',
            titulo='Informe MoreApp recibido',
            mensaje='Ya hay un registro MoreApp vinculado. Revise el detalle y cierre la comprobación.',
            accion_label='Ver informes MoreApp',
        )
    elif estado == 'REALIZADA_PENDIENTE_COMPROBACION':
        nivel = 'warning' if sync_advertencia else 'info'
        paso.update(
            nivel=nivel,
            titulo='Paso 4 — Comprobar informe de terreno',
            mensaje=(
                'MoreApp actualizó la orden. Revise equipos, advertencias y envíe a validación de Delco.'
                if not sync_advertencia
                else 'Hay advertencias en el informe MoreApp. Resuélvalas antes de validar.'
            ),
            accion_label='Revisar pendientes',
        )
    elif estado == 'PENDIENTE_VALIDACION':
        paso.update(
            nivel='warning',
            titulo='Paso 5 — Validación Delco',
            mensaje='La orden está lista para que el administrativo la valide.',
        )
    elif estado == 'VALIDADA':
        paso.update(
            nivel='success',
            titulo='Orden validada',
            mensaje='Trabajo aprobado. Use Acciones → Finalizada para cerrar el ciclo en la plataforma.',
            accion_label='Cerrar orden',
        )
    elif estado == 'OBSERVADA':
        paso.update(
            nivel='warning',
            titulo='Orden observada',
            mensaje='La validación rechazó el trabajo. Revise la OT derivada creada para el reintento en terreno.',
        )
    elif estado == 'FINALIZADA':
        paso.update(
            nivel='success',
            titulo='Orden finalizada',
            mensaje='Ciclo operativo cerrado.',
        )
    elif estado == 'CANCELADA':
        paso.update(
            nivel='secondary',
            titulo='Orden cancelada',
            mensaje='Esta orden no continúa en el flujo operativo.',
        )

    return paso


def crear_orden_derivada_por_observacion(
    orden: OrdenTrabajo,
    usuario: Usuario,
    observacion: str,
) -> OrdenTrabajo:
    """Crea una OT nueva vinculada cuando la validación marca OBSERVADA."""
    motivo = (observacion or '').strip()
    titulo_base = (orden.titulo or f'OT #{orden.pk}')[:150]
    nueva = OrdenTrabajo(
        titulo=f'OT derivada — {titulo_base}'[:200],
        descripcion=(
            f'Orden derivada de OT #{orden.pk} por observación en validación.\n'
            f'Motivo: {motivo}'
        ),
        tipo_trabajo=orden.tipo_trabajo,
        cliente=orden.cliente,
        tecnico_responsable=orden.tecnico_responsable,
        creada_por=usuario,
        orden_origen=orden,
        observaciones_tecnicas=f'Derivada de OT #{orden.pk}. Motivo: {motivo}',
        estado='ASIGNADA' if orden.tecnico_responsable_id else 'CREADA',
    )
    if nueva.estado == 'ASIGNADA':
        nueva.fecha_asignacion = timezone.now()
    nueva.save()
    aplicar_alerta_duplicado(nueva)
    return nueva
