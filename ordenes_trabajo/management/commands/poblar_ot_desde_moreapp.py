"""
Pobla clientes + OT desde IntegracionMoreApp y genera Excel de importación.

Los reportes operativos dependen de OT y de clientes cuyo numero_cliente
coincida con datos_procesados.cliente_codigo de MoreApp. Hoy suele haber
MoreApp procesado sin ficha ni OT → reportes en 0.

Uso:
  python manage.py poblar_ot_desde_moreapp --dry-run
  python manage.py poblar_ot_desde_moreapp --aplicar
  python manage.py poblar_ot_desde_moreapp --aplicar --excel
"""
from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date, parse_datetime

from clientes.models import Cliente
from ordenes_trabajo.models import IntegracionMoreApp, OrdenTrabajo
from ordenes_trabajo.sync import vincular_moreapp_a_orden
from usuarios.models import Usuario
from web.perf_cache import cache_invalidate


def _txt(valor: Any) -> str:
    if valor is None:
        return ''
    texto = str(valor).strip()
    if texto.endswith('.0') and texto[:-2].isdigit():
        return texto[:-2]
    return texto


def _tipo_trabajo(datos: dict, formulario: str) -> str:
    estado = _txt(datos.get('estado')).upper()
    if estado == 'INCIDENCIA':
        return 'INSPECCION'
    texto = f"{_txt(datos.get('trabajo'))} {_txt(datos.get('actividad'))}".upper()
    if 'RETIRO' in texto:
        return 'RETIRO'
    if 'CAMBIO' in texto:
        return 'CAMBIO'
    if 'REPROGRAM' in texto or 'MANTEN' in texto or 'REPAR' in texto:
        return 'MANTENCION'
    if 'Registro de Medidores' in (formulario or '') or 'TELEMEDIDA' in texto or 'INSTAL' in texto:
        return 'INSTALACION'
    if 'Mantenimiento' in (formulario or ''):
        return 'MANTENCION'
    return 'MANTENCION'


def _estado_ot(datos: dict) -> str:
    estado = _txt(datos.get('estado')).upper()
    if estado in {'EJECUTADO', 'REALIZADO', 'COMPLETADO', 'OK'}:
        return 'REALIZADA'
    if estado in {'INCIDENCIA', 'OBSERVADO'}:
        return 'OBSERVADA'
    return 'ASIGNADA'


def _buscar_tecnico(nombre: str) -> Optional[Usuario]:
    nombre = _txt(nombre)
    if not nombre:
        return None
    qs = Usuario.objects.filter(rol='TECNICO', is_active=True)
    exacto = qs.filter(nombre_interno__iexact=nombre).first()
    if exacto:
        return exacto
    return qs.filter(nombre_interno__icontains=nombre.split()[0]).first()


def _usuario_sistema() -> Usuario:
    admin = (
        Usuario.objects.filter(rol='ADMIN', is_active=True).order_by('id').first()
        or Usuario.objects.filter(is_superuser=True).order_by('id').first()
        or Usuario.objects.filter(is_active=True).order_by('id').first()
    )
    if not admin:
        raise RuntimeError('No hay usuario activo para crear OT (creada_por).')
    return admin


def _upsert_cliente(codigo: str, datos: dict) -> Tuple[Cliente, bool]:
    codigo = _txt(codigo)
    cliente = Cliente.objects.filter(numero_cliente=codigo).first()
    if not cliente:
        cliente = Cliente.objects.filter(numero_cliente__iexact=codigo).first()

    defaults = {
        'direccion': _txt(datos.get('cliente_direccion')) or f'Cliente {codigo}',
        'comuna': _txt(datos.get('cliente_comuna')) or 'Por definir',
        'customer_name': _txt(datos.get('cliente_nombre')) or None,
        'installation_address': _txt(datos.get('cliente_direccion')) or None,
        'meter_serial_n_1': _txt(datos.get('medidor_dejado_numero') or datos.get('serial_number')) or None,
        'meter_manufacturer_id': _txt(datos.get('marca_medidor_dejado') or datos.get('marca_aparato')) or None,
        'activo': True,
    }

    if cliente:
        changed = False
        for campo, valor in defaults.items():
            if valor and not getattr(cliente, campo, None):
                setattr(cliente, campo, valor)
                changed = True
        if changed:
            cliente.save()
        return cliente, False

    cliente = Cliente.objects.create(numero_cliente=codigo, **defaults)
    return cliente, True


def _parse_fecha_trabajo(datos: dict):
    raw = _txt(datos.get('fecha_trabajo') or datos.get('fecha_registro'))
    if not raw:
        return timezone.now()
    dt = parse_datetime(raw)
    if dt:
        if timezone.is_naive(dt):
            dt = timezone.make_aware(dt, timezone.get_current_timezone())
        return dt
    d = parse_date(raw[:10])
    if d:
        return timezone.make_aware(datetime.combine(d, datetime.min.time()))
    return timezone.now()


def _iter_registros_moreapp():
    return (
        IntegracionMoreApp.objects.filter(eliminado=False)
        .exclude(datos_procesados__cliente_codigo__isnull=True)
        .exclude(datos_procesados__cliente_codigo='')
        .order_by('numero_correlativo', 'id')
    )


class Command(BaseCommand):
    help = 'Crea/actualiza clientes y OT desde MoreApp; opcionalmente genera Excel de importación.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Solo muestra el plan (equivalente a no pasar --aplicar/--excel).',
        )
        parser.add_argument(
            '--excel',
            action='store_true',
            help='Genera datos_prueba/ordenes_import_moreapp.xlsx para importar por UI.',
        )
        parser.add_argument(
            '--una-por-cliente',
            action='store_true',
            default=True,
            help='Una sola OT por Nº cliente (la de correlativo más alto). Default: sí.',
        )
        parser.add_argument(
            '--todas',
            action='store_true',
            help='Crea una OT por cada registro MoreApp (puede duplicar clientes).',
        )

    def handle(self, *args, **options):
        aplicar = options['aplicar']
        generar_excel = options['excel']
        una_por_cliente = not options['todas']

        registros = list(_iter_registros_moreapp())
        if not registros:
            self.stderr.write(self.style.ERROR('No hay registros MoreApp con cliente_codigo.'))
            return

        # Agrupar por cliente; si una_por_cliente, quedarse con el correlativo mayor
        por_cliente: Dict[str, List[IntegracionMoreApp]] = defaultdict(list)
        for reg in registros:
            datos = reg.datos_procesados or {}
            codigo = _txt(datos.get('cliente_codigo'))
            if codigo:
                por_cliente[codigo].append(reg)

        seleccionados: List[IntegracionMoreApp] = []
        for codigo, items in por_cliente.items():
            items.sort(
                key=lambda r: (
                    int(r.numero_correlativo) if r.numero_correlativo is not None else 0,
                    r.pk,
                ),
                reverse=True,
            )
            if una_por_cliente:
                seleccionados.append(items[0])
            else:
                seleccionados.extend(items)

        self.stdout.write(
            f'MoreApp con código: {len(registros)} | '
            f'Clientes únicos: {len(por_cliente)} | '
            f'OT a procesar: {len(seleccionados)}'
        )

        existentes = Cliente.objects.filter(
            numero_cliente__in=list(por_cliente.keys())
        ).count()
        self.stdout.write(
            f'Clientes ya en BD: {existentes} | Faltan: {len(por_cliente) - existentes}'
        )
        self.stdout.write(f'OT actuales en BD: {OrdenTrabajo.objects.count()}')

        if not aplicar and not generar_excel:
            self.stdout.write(self.style.WARNING(
                'Dry-run. Usa --aplicar para crear clientes/OT, y/o --excel para el archivo.'
            ))
            for reg in seleccionados[:12]:
                d = reg.datos_procesados or {}
                self.stdout.write(
                    f"  corr={reg.numero_correlativo} cliente={d.get('cliente_codigo')} "
                    f"estado_ma={d.get('estado')} -> OT {_estado_ot(d)} | "
                    f"{_txt(d.get('cliente_nombre'))[:40]}"
                )
            if len(seleccionados) > 12:
                self.stdout.write(f'  ... y {len(seleccionados) - 12} más')
            return

        filas_excel: List[list] = []
        stats = {
            'clientes_creados': 0,
            'clientes_reusados': 0,
            'ot_creadas': 0,
            'ot_reusadas': 0,
            'vinculos': 0,
            'errores': 0,
        }

        if aplicar:
            creador = _usuario_sistema()
            with transaction.atomic():
                for reg in seleccionados:
                    try:
                        datos = reg.datos_procesados or {}
                        codigo = _txt(datos.get('cliente_codigo'))
                        cliente, creado = _upsert_cliente(codigo, datos)
                        if creado:
                            stats['clientes_creados'] += 1
                        else:
                            stats['clientes_reusados'] += 1

                        tipo = _tipo_trabajo(datos, reg.nombre_formulario or '')
                        estado = _estado_ot(datos)
                        tecnico = _buscar_tecnico(datos.get('tecnico_responsable'))
                        fecha = _parse_fecha_trabajo(datos)
                        titulo = (
                            f"OT MoreApp #{reg.numero_correlativo or reg.pk} — "
                            f"{_txt(datos.get('cliente_nombre')) or codigo}"
                        )[:200]
                        descripcion = (
                            f"{reg.nombre_formulario or 'MoreApp'} | "
                            f"Correlativo {reg.numero_correlativo or '—'} | "
                            f"Estado terreno: {_txt(datos.get('estado')) or '—'}"
                        )
                        if _txt(datos.get('trabajo') or datos.get('actividad')):
                            descripcion += f" | {_txt(datos.get('trabajo') or datos.get('actividad'))}"

                        # Reutilizar OT abierta del mismo cliente+tipo si existe
                        orden = (
                            OrdenTrabajo.objects.filter(
                                cliente=cliente,
                                tipo_trabajo=tipo,
                                estado__in=list(OrdenTrabajo.ESTADOS_ABIERTOS) + [
                                    'REALIZADA', 'OBSERVADA', 'VALIDADA', 'FINALIZADA',
                                ],
                            )
                            .order_by('-fecha_creacion')
                            .first()
                        )

                        if orden and una_por_cliente:
                            stats['ot_reusadas'] += 1
                            # Si MoreApp dice ejecutado y la OT sigue abierta, cerrarla
                            if estado == 'REALIZADA' and orden.estado in OrdenTrabajo.ESTADOS_ABIERTOS:
                                orden.estado = 'REALIZADA'
                                if not orden.fecha_inicio_ejecucion:
                                    orden.fecha_inicio_ejecucion = fecha
                                if not orden.fecha_finalizacion:
                                    orden.fecha_finalizacion = fecha
                                orden.save()
                        else:
                            orden = OrdenTrabajo(
                                titulo=titulo,
                                descripcion=descripcion[:2000],
                                tipo_trabajo=tipo,
                                cliente=cliente,
                                estado=estado,
                                tecnico_responsable=tecnico,
                                creada_por=creador,
                                observaciones_tecnicas=(
                                    f"MoreApp submission: {reg.moreapp_submission_id} | "
                                    f"correlativo: {reg.numero_correlativo}"
                                )[:2000],
                            )
                            if estado in {'ASIGNADA', 'EN_EJECUCION'} and tecnico:
                                orden.fecha_asignacion = fecha
                            if estado == 'REALIZADA':
                                orden.fecha_asignacion = fecha
                                orden.fecha_inicio_ejecucion = fecha
                                orden.fecha_finalizacion = fecha
                            orden.save()
                            stats['ot_creadas'] += 1

                        vinculada = vincular_moreapp_a_orden(
                            cliente=cliente,
                            registro_moreapp=reg,
                            usuario=creador,
                        )
                        if vinculada or reg.orden_id:
                            stats['vinculos'] += 1
                            if not reg.orden_id and orden:
                                reg.orden = orden
                                reg.save(update_fields=['orden'])

                        filas_excel.append([
                            codigo,
                            titulo,
                            descripcion[:500],
                            tipo,
                            _txt(datos.get('tecnico_responsable')),
                            estado,
                            _txt(datos.get('cliente_direccion')) or f'Direccion cliente {codigo}',
                            _txt(datos.get('cliente_comuna')) or 'Por definir',
                            f'MoreApp correlativo {reg.numero_correlativo}',
                            reg.numero_correlativo or '',
                            reg.nombre_formulario or '',
                        ])
                    except Exception as exc:
                        stats['errores'] += 1
                        self.stderr.write(self.style.ERROR(
                            f'Error corr={getattr(reg, "numero_correlativo", "?")}: {exc}'
                        ))

            cache_invalidate('operacional:codigos_moreapp', 'operacional:cliente_ids', 'moreapp:aviso_conteos')
            self.stdout.write(self.style.SUCCESS(
                f"Aplicado: clientes +{stats['clientes_creados']} "
                f"(reusados {stats['clientes_reusados']}), "
                f"OT +{stats['ot_creadas']} (reusadas {stats['ot_reusadas']}), "
                f"vínculos {stats['vinculos']}, errores {stats['errores']}"
            ))
            self.stdout.write(
                f'Totales ahora → Clientes operativos potenciales: {len(por_cliente)} | '
                f'OT: {OrdenTrabajo.objects.count()} | '
                f'MoreApp con OT: {IntegracionMoreApp.objects.filter(eliminado=False, orden__isnull=False).count()}'
            )

        if generar_excel:
            if not filas_excel:
                # Dry excel from selección
                for reg in seleccionados:
                    datos = reg.datos_procesados or {}
                    codigo = _txt(datos.get('cliente_codigo'))
                    filas_excel.append([
                        codigo,
                        f"OT MoreApp #{reg.numero_correlativo or reg.pk} — {_txt(datos.get('cliente_nombre')) or codigo}"[:200],
                        f"{reg.nombre_formulario or 'MoreApp'} | Correlativo {reg.numero_correlativo}"[:500],
                        _tipo_trabajo(datos, reg.nombre_formulario or ''),
                        _txt(datos.get('tecnico_responsable')),
                        _estado_ot(datos),
                        _txt(datos.get('cliente_direccion')) or f'Direccion cliente {codigo}',
                        _txt(datos.get('cliente_comuna')) or 'Por definir',
                        f'MoreApp correlativo {reg.numero_correlativo}',
                        reg.numero_correlativo or '',
                        reg.nombre_formulario or '',
                    ])
            path = self._guardar_excel(filas_excel)
            self.stdout.write(self.style.SUCCESS(f'Excel generado: {path}'))

    def _guardar_excel(self, filas: List[list]) -> str:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        headers = [
            'Numero Cliente',
            'Titulo',
            'Descripcion',
            'Tipo Trabajo',
            'Tecnico Responsable',
            'Estado',
            'Direccion Cliente',
            'Comuna',
            'Observaciones Tecnicas',
            'Correlativo MoreApp',
            'Formulario MoreApp',
        ]
        out_dir = Path(__file__).resolve().parents[3] / 'datos_prueba'
        # parents: commands -> management -> ordenes_trabajo -> project root
        out_dir = Path(__file__).resolve().parents[3]
        # Actually: .../ordenes_trabajo/management/commands/this.py
        # parents[0]=commands, [1]=management, [2]=ordenes_trabajo, [3]=project root
        proyecto = Path(__file__).resolve().parents[3]
        out_dir = proyecto / 'datos_prueba'
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / 'ordenes_import_moreapp.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Una OT por cliente'
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for fila in filas:
            ws.append(fila)

        inst = wb.create_sheet('Instrucciones')
        for line in [
            ['Importar OT desde MoreApp'],
            ['1. Ir a Órdenes de trabajo → Importar Excel'],
            ['2. Usar la hoja "Una OT por cliente"'],
            ['3. Si el Nº cliente no existe, el import lo crea'],
            ['4. Preferible: python manage.py poblar_ot_desde_moreapp --aplicar'],
            [f'Filas en este archivo: {len(filas)}'],
            [f'Generado: {timezone.now().isoformat()}'],
        ]:
            inst.append(line)

        wb.save(out_path)
        return str(out_path)
