"""
Utilidades para importación desde Excel
"""

from __future__ import annotations

import json
import logging
import re
import unicodedata

import openpyxl
from importaciones.models import ImportacionExcel, ImportacionExcelError
from inventario.models import Medidor, SimCard, Modem, EstadoInventario, Ubicacion, MovimientoInventario, MovimientoItem
from clientes.models import Cliente
from usuarios.models import Usuario
from web.services.validators import (
    merge_issues,
    normalize_ip_value,
    validate_ip_format,
    validate_ip_port_coherence,
    validate_meter_uniqueness,
    validate_modem_assignment,
)
from web.services.audit import AuditEvent, register_audit_event


from django.db import OperationalError, transaction
from django.db.models import Q
from django.db.utils import IntegrityError

logger = logging.getLogger(__name__)


def aplicar_estilo_hoja_exportacion(
    ws,
    *,
    header_fill_hex: str = '1F4E79',
    freeze: str = 'A2',
    zebra: bool = True,
    auto_filter: bool = False,
    filter_from_col: int = 1,
    filter_to_col: int = None,
    min_width: int = 10,
    max_width: int = 40,
    header_height: int = 30,
    header_row: int = 1,
):
    """Aplica formato visual uniforme a una hoja de exportación Excel.

    No altera valores ni encabezados: solo presentación para revisión cómoda.

    auto_filter: por defecto False (muchas columnas no aportan filtro útil).
    filter_from_col / filter_to_col: rango 1-based de columnas con filtro
    (p. ej. saltar '#' o textos largos / IDs).
    header_row: fila 1-based donde están los encabezados (útil si hay título arriba).
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    if ws.max_row < 1 or ws.max_column < 1:
        return ws

    header_row = max(1, int(header_row or 1))
    if header_row > ws.max_row:
        return ws

    header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    body_font = Font(size=10, name='Calibri', color='212121')
    body_align = Alignment(vertical='center', wrap_text=True)
    zebra_fill = PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid')
    thin = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE'),
    )

    ws.row_dimensions[header_row].height = header_height
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=ws.max_column):
        ws.row_dimensions[row[0].row].height = 18
        for cell in row:
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin
            if zebra and ((cell.row - header_row) % 2 == 0):
                cell.fill = zebra_fill

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            if cell.row < header_row:
                continue
            if cell.value is None:
                continue
            texto = str(cell.value).split('\n', 1)[0].strip()
            if len(texto) > max_len:
                max_len = len(texto)
        ws.column_dimensions[letter].width = min(max(max_len + 3, min_width), max_width)

    if freeze:
        # Si el freeze por defecto era A2 y movimos el encabezado, ajustar.
        if freeze == 'A2' and header_row != 1:
            freeze = f'A{header_row + 1}'
        ws.freeze_panes = freeze
    if auto_filter and ws.max_row >= header_row:
        inicio = max(1, min(int(filter_from_col or 1), ws.max_column))
        fin = ws.max_column if filter_to_col is None else int(filter_to_col)
        fin = max(inicio, min(fin, ws.max_column))
        ws.auto_filter.ref = (
            f'{get_column_letter(inicio)}{header_row}:'
            f'{get_column_letter(fin)}{ws.max_row}'
        )

    ws.sheet_view.showGridLines = False
    return ws


# Encabezados legibles para exportación completa (evita cliente_label, nombre_interno, etc.)
_ETIQUETAS_EXPORT_COMPLETO = {
    'id': 'ID',
    'serie': 'Serie',
    'imei': 'IMEI',
    'marca': 'Marca',
    'modelo': 'Modelo',
    'caja': 'Caja',
    'bodega': 'Bodega',
    'modulo': 'Módulo',
    'tipo_medidor': 'Tipo medidor',
    'tipo_medidor_display': 'Tipo medidor (texto)',
    'fecha_recepcion': 'Fecha recepción',
    'fecha_entrega': 'Fecha entrega',
    'fecha_creacion': 'Fecha creación',
    'fecha_actualizacion': 'Fecha actualización',
    'fecha_eliminacion': 'Fecha eliminación',
    'entregado_a_info': 'Entregado a (texto Excel)',
    'entregado_a_otro': 'Entregado a (manual)',
    'entregado_a_nombre': 'Entregado a',
    'entregado_a_id': 'ID entregado a',
    'entregado_a_label': 'Entregado a',
    'en_custodia_de_id': 'ID en custodia de',
    'en_custodia_de_label': 'En custodia de',
    'estado_inventario_id': 'ID estado',
    'estado_inventario_label': 'Estado',
    'ubicacion_actual_id': 'ID ubicación',
    'ubicacion_actual_label': 'Ubicación',
    'cliente_id': 'ID cliente',
    'cliente_label': 'Nº cliente',
    'cliente_otro': 'Cliente (manual)',
    'medidor_id': 'ID medidor',
    'medidor_label': 'Serie medidor',
    'medidor_otro': 'Medidor (manual)',
    'eliminado_por_id': 'ID eliminado por',
    'eliminado_por_label': 'Eliminado por',
    'eliminado': 'Eliminado',
    'proyecto': 'Proyecto',
    'observaciones': 'Observaciones',
    'operador': 'Operador',
    'abonado': 'Abonado',
    'direccion_ip': 'Dirección IP',
    'apn': 'APN',
    'msisdn': 'MSISDN',
    'ip_fija': 'IP fija',
    'serie_plastico': 'Serie plástico',
    'proveedor': 'Proveedor',
    'tecnico_responsable': 'Técnico responsable',
    'ip': 'IP',
    'puerto': 'Puerto',
    'marca_secundaria': 'Marca secundaria',
    'retirado': 'Retirado',
    'serie_secundaria': 'Serie secundaria',
    'irregularidad': 'Irregularidad',
    # Cliente: campos con nombre técnico en inglés → encabezado en español
    'numero_cliente': 'Número cliente',
    'direccion': 'Dirección',
    'customer_name': 'Nombre cliente',
    'installation_address': 'Dirección instalación',
    'city': 'Ciudad',
    'client_type': 'Tipo cliente',
    'note': 'Nota',
    'meter_manufacturer_id': 'Marca medidor',
    'meter_serial_n_1': 'Serie medidor',
    'medidor_actual_id': 'ID medidor actual',
    'medidor_actual_label': 'Medidor actual (serie)',
    'sim_operador': 'SIM operador',
    'sim_iccid': 'SIM ICCID',
    'sim_abonado': 'SIM abonado',
    'sim_estado': 'SIM estado',
    'sim_estado_display': 'SIM estado (texto)',
    'estado_telemetria': 'Estado telemetría',
    'estado_telemetria_display': 'Estado telemetría (texto)',
    'estado_stb': 'Estado STB',
    'estado_stb_display': 'Estado STB (texto)',
    'estado_sci4': 'Estado SCi4',
    'estado_sci4_display': 'Estado SCi4 (texto)',
    'ultimo_acceso': 'Último acceso',
    'ultimo_perfil_carga': 'Último perfil carga',
    'ultimo_perfil_instrumentacion': 'Último perfil instrumentación',
    'ultimo_reset': 'Último reset',
    'ultimo_registro_facturacion': 'Último registro facturación',
    'fecha_registro': 'Fecha registro',
}


def _etiqueta_columna_export_completo(clave: str) -> str:
    """Convierte claves técnicas (snake_case) en encabezados legibles."""
    if clave in _ETIQUETAS_EXPORT_COMPLETO:
        return _ETIQUETAS_EXPORT_COMPLETO[clave]

    sufijo = ''
    base = clave
    if clave.endswith('_display'):
        base, sufijo = clave[:-8], ' (texto)'
    elif clave.endswith('_label'):
        base = clave[:-6]
    elif clave.endswith('_id'):
        base, sufijo = clave[:-3], ' (ID)'

    texto = base.replace('_', ' ').strip()
    if not texto:
        return clave
    return texto[0].upper() + texto[1:] + sufijo


def _normalizar_serie_medidor_cliente(valor) -> str:
    """Normaliza serie de medidor para comparar fichas de cliente."""
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    if texto.endswith('.0') and texto[:-2].isdigit():
        return texto[:-2]
    return texto


def _serie_cliente_vacia(serie) -> bool:
    return not _normalizar_serie_medidor_cliente(serie)


def _serie_cliente_igual(a, b) -> bool:
    return _normalizar_serie_medidor_cliente(a).casefold() == _normalizar_serie_medidor_cliente(b).casefold()


def _clave_cliente_sync(numero, serie) -> tuple:
    """Clave estable para sincronización completa (Nº + serie normalizada)."""
    return (
        str(numero or '').strip(),
        _normalizar_serie_medidor_cliente(serie).casefold(),
    )


def _buscar_cliente_para_importacion(numero_text: str, serie_text: str):
    """Resuelve ficha a actualizar/reactivar para evitar duplicados innecesarios.

    Orden:
    1) Match exacto Nº + serie (activo, luego inactivo).
    2) Una sola ficha activa de ese Nº → actualizarla (aunque cambie la serie).
    3) Match por serie entre activas del mismo Nº.
    4) Igual lógica sobre inactivas (reactivar).
    5) None → crear (p. ej. mismo Nº con otra serie cuando ya hay varias fichas).
    """
    serie_text = _normalizar_serie_medidor_cliente(serie_text)
    base = Cliente.objects.filter(numero_cliente=numero_text)

    def _match_serie(qs):
        if serie_text:
            for obj in qs:
                if _serie_cliente_igual(obj.meter_serial_n_1, serie_text):
                    return obj
            return None
        for obj in qs:
            if _serie_cliente_vacia(obj.meter_serial_n_1):
                return obj
        return None

    activos = list(base.filter(activo=True).order_by('id'))
    exacto_activo = _match_serie(activos)
    if exacto_activo:
        return exacto_activo, False

    if len(activos) == 1:
        actual = activos[0]
        actual_serie = _normalizar_serie_medidor_cliente(actual.meter_serial_n_1)
        # Misma ficha si no hay serie nueva, la actual no tiene serie, o coinciden.
        if not serie_text or not actual_serie or _serie_cliente_igual(actual_serie, serie_text):
            return actual, False
        # Serie distinta y ambas informadas → otro medidor del mismo Nº (crear).

    inactivos = list(base.filter(activo=False).order_by('-fecha_actualizacion', '-id'))
    exacto_inactivo = _match_serie(inactivos)
    if exacto_inactivo:
        return exacto_inactivo, True

    if not activos and len(inactivos) == 1:
        actual = inactivos[0]
        actual_serie = _normalizar_serie_medidor_cliente(actual.meter_serial_n_1)
        if not serie_text or not actual_serie or _serie_cliente_igual(actual_serie, serie_text):
            return actual, True

    if not activos and inactivos and not serie_text:
        vacio = next((c for c in inactivos if _serie_cliente_vacia(c.meter_serial_n_1)), None)
        return (vacio or inactivos[0]), True

    return None, False


def _con_reintento_sqlite(callable_fn, intentos=6, espera=0.4):
    """Reintenta operaciones cuando SQLite responde database is locked."""
    import time

    ultimo = None
    for i in range(intentos):
        try:
            return callable_fn()
        except OperationalError as exc:
            ultimo = exc
            if 'locked' not in str(exc).lower() or i == intentos - 1:
                raise
            time.sleep(espera * (i + 1))
    raise ultimo


def _normalizar_header_equipo(valor) -> str:
    if valor is None:
        return ''
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
    texto = texto.replace(':', '').replace('#', 'numero')
    texto = texto.replace(' ', '_')
    while '__' in texto:
        texto = texto.replace('__', '_')
    return texto


def _seleccionar_hoja_equipos(wb, tipo_equipo: str):
    """Elige la hoja de datos real (evita pivotes tipo Consolidado)."""
    tipo = (tipo_equipo or '').upper()
    preferidas = {
        'MEDIDORES': ('medidor', 'directo', 'indirecto'),
        'SIM': ('sim',),
        'MODEMS': ('modem', 'modem', 'maestro modem'),
    }
    excluidas = ('consolidado', 'hoja1', 'hoja2', 'hoja3', 'hoja4', 'resumen', 'pivot')

    nombres = list(wb.sheetnames)
    if len(nombres) == 1:
        return wb[nombres[0]]

    claves = preferidas.get(tipo, ())
    for nombre in nombres:
        n = _normalizar_header_equipo(nombre).replace('_', ' ')
        if any(n.startswith(ex) or n == ex for ex in excluidas):
            continue
        if any(clave in n for clave in claves):
            return wb[nombre]

    # Fallback: primera hoja que no sea consolidado/hoja auxiliar
    for nombre in nombres:
        n = _normalizar_header_equipo(nombre).replace('_', ' ')
        if not any(n.startswith(ex) or n == ex for ex in excluidas):
            return wb[nombre]
    return wb.active


def _resolver_estado_inventario(nombre_estado, default_nombre='En bodega'):
    """Mapea estados del Excel a EstadoInventario del sistema."""
    if nombre_estado is None or str(nombre_estado).strip() == '':
        estado = EstadoInventario.objects.filter(nombre=default_nombre).first()
        if not estado:
            estado = EstadoInventario.objects.create(nombre=default_nombre)
        return estado

    texto = str(nombre_estado).strip()
    texto_n = _normalizar_header_equipo(texto).replace('_', ' ')

    aliases = {
        'disponible': 'En bodega',
        'bodega': 'En bodega',
        'en bodega': 'En bodega',
        'instalado': 'Instalado',
        'trayecto': 'En Trayecto',
        'en trayecto': 'En Trayecto',
        'retirado': 'Retirado',
        'devuelta': 'Devuelta',
        'devuelto': 'Devuelta',
        'sin conexion': 'Sin Conexión',
        'con problemas': 'Con Problemas',
        'mal estado': 'Dado de baja',
        'custodia': 'En custodia técnico',
        'en custodia tecnico': 'En custodia técnico',
        'peaje': 'En peaje',
        'en peaje': 'En peaje',
        'reparacion': 'En reparación',
        'en reparacion': 'En reparación',
        'revision': 'En revisión',
        'en revision': 'En revisión',
        'dado de baja': 'Dado de baja',
        'malo': 'Dado de baja',
    }
    for clave, destino in aliases.items():
        if clave in texto_n:
            estado = EstadoInventario.objects.filter(nombre=destino).first()
            if not estado:
                estado = EstadoInventario.objects.create(nombre=destino)
            return estado

    estado = EstadoInventario.objects.filter(nombre__iexact=texto).first()
    if estado:
        return estado
    estado = EstadoInventario.objects.filter(nombre__icontains=texto).first()
    if estado:
        return estado

    estado = EstadoInventario.objects.filter(nombre=default_nombre).first()
    if not estado:
        estado = EstadoInventario.objects.create(nombre=default_nombre)
    return estado


def _parse_fecha_flexible(valor, nombre_columna='Fecha'):
    from datetime import datetime, date
    if valor is None or valor == '':
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if hasattr(valor, 'date') and not isinstance(valor, (str, bytes)):
        try:
            return valor.date()
        except Exception:
            pass

    texto = str(valor).strip()
    if 'datetime.datetime' in texto:
        import re
        match = re.search(r'datetime\.datetime\s*\(\s*(\d+),\s*(\d+),\s*(\d+)', texto)
        if match:
            y, m, d = match.groups()
            return date(int(y), int(m), int(d))

    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(texto, fmt).date()
        except Exception:
            continue
    raise ValueError(f'Columna "{nombre_columna}": formato no válido. Valor recibido: {valor}')


def _as_text_id(valor) -> str:
    if valor is None:
        return ''
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    if isinstance(valor, int):
        return str(valor)
    return str(valor).strip().strip("'").strip('"').strip()


def _normalizar_caja_modem(valor) -> str:
    """
    Conserva el texto de caja del Excel (ej. 'Caja 1', 'Caja 260318 014').
    Solo limpia espacios / basura; no se altera el número de lote.
    """
    return _as_text_id(valor)


def _reconstruir_ipv4_desde_digitos(digits: str):
    """Reconstruye candidatos IPv4 cuando Excel guardó la IP como entero sin puntos."""
    n = len(digits)
    candidatos = []
    for a in range(1, 4):
        for b in range(1, 4):
            for c in range(1, 4):
                d = n - a - b - c
                if d < 1 or d > 3:
                    continue
                parts = [digits[0:a], digits[a:a + b], digits[a + b:a + b + c], digits[a + b + c:]]
                if any(len(p) > 1 and p[0] == '0' for p in parts):
                    continue
                octets = [int(p) for p in parts]
                if all(0 <= o <= 255 for o in octets):
                    candidatos.append('.'.join(str(o) for o in octets))
    vistos = set()
    unicos = []
    for ip in candidatos:
        if ip not in vistos:
            vistos.add(ip)
            unicos.append(ip)
    return unicos


def _normalizar_direccion_ip(valor) -> str:
    """
    Normaliza IP desde Excel.
    Excel a veces convierte 10.118.170.179 en el entero 10118170179.
    """
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    texto = str(valor).strip().strip("'").strip('"')
    if not texto or texto.upper() in {'NONE', 'NULL', 'N/A', '#N/A'}:
        return ''

    if re.fullmatch(r'\d{1,3}(\.\d{1,3}){3}', texto):
        partes = [int(p) for p in texto.split('.')]
        if all(0 <= p <= 255 for p in partes):
            return '.'.join(str(p) for p in partes)
        return texto

    digits = re.sub(r'\D', '', texto)
    if re.fullmatch(r'\d{8,12}', digits) and '.' not in texto:
        candidatos = _reconstruir_ipv4_desde_digitos(digits)
        if not candidatos:
            return texto
        preferidos = [ip for ip in candidatos if ip.startswith('10.118.')]
        if not preferidos:
            preferidos = [ip for ip in candidatos if ip.startswith('10.')]
        if not preferidos:
            preferidos = candidatos
        preferidos.sort(
            key=lambda ip: (len(ip.split('.')[2]), int(ip.split('.')[2])),
            reverse=True,
        )
        return preferidos[0]

    return texto


def importar_equipos_excel(archivo, usuario, tipo_equipo='MEDIDORES'):
    """
    Importa equipos (Medidores, SIM, Módems) desde archivo Excel.

    - Elige la hoja correcta (no pivotes).
    - Mapea columnas por encabezado.
    - Si el equipo ya existe (serie/IMEI), ACTUALIZA en vez de duplicar.
    """
    def error_columna(columna, mensaje, valor=None):
        base = f'Columna "{columna}": {mensaje}'
        if valor is not None and str(valor).strip() != '':
            return f'{base}. Valor recibido: {valor}'
        return base

    def fila_a_texto(headers_fila, valores_fila):
        data = {}
        for i, valor in enumerate(valores_fila):
            nombre_columna = headers_fila[i] if i < len(headers_fila) else None
            nombre_columna = str(nombre_columna).strip() if nombre_columna else f'Columna_{i + 1}'
            data[nombre_columna] = valor
        return json.dumps(data, ensure_ascii=False, default=str)

    def registrar_movimiento_importacion(equipo_obj, tipo_item, detalle, fila):
        try:
            movimiento = MovimientoInventario.objects.create(
                tipo='IMPORTACION',
                origen=bodega,
                destino=bodega,
                responsable=usuario,
                observacion=f'Importación masiva {tipo_item} fila {fila}: {detalle}',
            )
            item_kwargs = {
                'movimiento': movimiento,
                'tipo_equipo': tipo_item,
                'cantidad': 1,
            }
            if tipo_item == 'MEDIDOR':
                item_kwargs['medidor'] = equipo_obj
            elif tipo_item == 'SIM':
                item_kwargs['simcard'] = equipo_obj
            else:
                item_kwargs['modem'] = equipo_obj
            MovimientoItem.objects.create(**item_kwargs)
        except Exception:
            pass

    def obtener_o_crear_cliente(numero, cache_clientes):
        if numero is None or str(numero).strip() == '':
            return None
        num = _as_text_id(numero)
        if not num or num.upper() in {'#N/A', 'N/A', 'NONE', 'NULL'}:
            return None
        if num in cache_clientes:
            return cache_clientes[num]

        def _get():
            cliente = Cliente.objects.filter(numero_cliente=num, activo=True).first()
            if cliente:
                return cliente
            inactivo = Cliente.objects.filter(numero_cliente=num, activo=False).first()
            if inactivo:
                inactivo.activo = True
                inactivo.fecha_eliminacion = None
                inactivo.eliminado_por = None
                inactivo.save(update_fields=['activo', 'fecha_eliminacion', 'eliminado_por'])
                return inactivo
            return Cliente.objects.create(
                numero_cliente=num,
                direccion=f'Cliente {num}',
                comuna='Por definir',
                activo=True,
            )

        cliente = _con_reintento_sqlite(_get)
        cache_clientes[num] = cliente
        return cliente

    importacion = ImportacionExcel.objects.create(
        tipo='EQUIPOS',
        archivo_original=archivo.name if hasattr(archivo, 'name') else 'Upload',
        usuario=usuario,
    )

    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = _seleccionar_hoja_equipos(wb, tipo_equipo)

        headers = [cell.value for cell in ws[1]]
        headers_norm = [_normalizar_header_equipo(h) for h in headers]
        logger.debug('Hoja equipos=%s headers=%s', ws.title, headers)

        def idx_col(*aliases):
            for alias in aliases:
                alias_n = _normalizar_header_equipo(alias)
                if alias_n in headers_norm:
                    return headers_norm.index(alias_n)
            return None

        def get_val(valores, *aliases):
            i = idx_col(*aliases)
            if i is None or i >= len(valores):
                return None
            return valores[i]

        bodega = Ubicacion.objects.filter(nombre__icontains='Bodega').first()
        if not bodega:
            bodega = Ubicacion.objects.create(tipo='BODEGA_DELCO', nombre='Bodega Principal')

        contador_filas = 0
        exitosas = 0
        fallidas = 0
        creadas = 0
        actualizadas = 0
        tipo = tipo_equipo.upper()
        cache_clientes = {}
        cache_medidores = {}
        # En Maestro Módem, "Caja N" solo aparece en la 1ª fila del grupo (~5 equipos);
        # las siguientes vienen vacías y heredan la misma caja.
        caja_modem_actual = ''
        imeis_modem_en_import = {}

        def buscar_medidor(serie_medidor):
            serie_m = _as_text_id(serie_medidor)
            if not serie_m:
                return None
            if serie_m in cache_medidores:
                return cache_medidores[serie_m]
            obj = Medidor.objects.filter(serie=serie_m, eliminado=False).first()
            cache_medidores[serie_m] = obj
            return obj

        avisos_reactivacion = []

        def _upsert_equipo(model, lookup_field, lookup_value, defaults):
            """Actualiza/crea solo equipos activos; revive soft-deleted de forma explícita.

            Returns: (objeto, creado, reactivado)
            """
            activo = model.objects.filter(**{lookup_field: lookup_value}, eliminado=False).first()
            if activo:
                for key, value in defaults.items():
                    setattr(activo, key, value)
                activo.save()
                return activo, False, False

            eliminado = model.objects.filter(**{lookup_field: lookup_value}, eliminado=True).first()
            if eliminado:
                for key, value in defaults.items():
                    setattr(eliminado, key, value)
                eliminado.eliminado = False
                if hasattr(eliminado, 'fecha_eliminacion'):
                    eliminado.fecha_eliminacion = None
                if hasattr(eliminado, 'eliminado_por_id'):
                    eliminado.eliminado_por = None
                eliminado.save()
                return eliminado, False, True

            return model.objects.create(**{lookup_field: lookup_value}, **defaults), True, False

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            valores = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != '' for v in valores):
                continue

            contador_filas += 1
            try:
                if tipo == 'MEDIDORES':
                    fecha_recepcion_raw = get_val(valores, 'fecha_recepcion', 'fecha_de_recepcion')
                    serie_raw = get_val(valores, 'medidor', 'serie')
                    tipo_medidor_raw = get_val(valores, 'tipo_medidor', 'tipo_de_medidor')
                    if not any([fecha_recepcion_raw, serie_raw, tipo_medidor_raw]):
                        contador_filas -= 1
                        continue

                    faltantes = []
                    if not fecha_recepcion_raw:
                        faltantes.append('Fecha Recepción')
                    if not serie_raw:
                        faltantes.append('Medidor/Serie')
                    if not tipo_medidor_raw:
                        faltantes.append('Tipo Medidor')
                    if faltantes:
                        raise ValueError(f'Faltan columnas requeridas: {", ".join(faltantes)}')

                    serie = _as_text_id(serie_raw)
                    marca = _as_text_id(get_val(valores, 'marca')) or None
                    caja = _as_text_id(get_val(valores, 'caja')) or None
                    tipo_medidor = _as_text_id(tipo_medidor_raw).upper()
                    if tipo_medidor in {'D', 'DIR', 'DIRECTO'}:
                        tipo_medidor = 'DIRECTO'
                    elif tipo_medidor in {'I', 'IND', 'INDIRECTO'}:
                        tipo_medidor = 'INDIRECTO'
                    else:
                        raise ValueError(error_columna('Tipo Medidor', 'valores válidos: DIRECTO o INDIRECTO', tipo_medidor))

                    fecha_recepcion = _parse_fecha_flexible(fecha_recepcion_raw, 'Fecha Recepción')
                    fecha_entrega = _parse_fecha_flexible(get_val(valores, 'fecha_de_entrega', 'fecha_entrega'), 'Fecha Entrega')

                    modulo = get_val(valores, 'modulo')
                    if isinstance(modulo, str):
                        m = modulo.strip().lower()
                        modulo = True if m in {'si', 'sí', 'yes', 'true', '1'} else False if m in {'no', 'false', '0'} else None
                    elif isinstance(modulo, (int, bool)):
                        modulo = bool(modulo)
                    else:
                        modulo = None

                    estado_obj = _resolver_estado_inventario(get_val(valores, 'estado'), 'En bodega')
                    cliente_raw = get_val(valores, 'cliente')
                    cliente_obj = obtener_o_crear_cliente(cliente_raw, cache_clientes)
                    cliente_texto = _as_text_id(cliente_raw)
                    if cliente_texto.upper() in {'#N/A', 'N/A', 'NONE', 'NULL'}:
                        cliente_texto = ''
                    entregado_a_info = _as_text_id(get_val(valores, 'entregado_a', 'entregado_a_'))
                    bodega_ref = _as_text_id(get_val(valores, 'bodega'))
                    correlativo = get_val(valores, 'numero', '#')

                    defaults = {
                        'fecha_recepcion': fecha_recepcion,
                        'bodega': bodega_ref,
                        'marca': marca or '',
                        'caja': caja,
                        'modulo': modulo,
                        'tipo_medidor': tipo_medidor,
                        'fecha_entrega': fecha_entrega,
                        'estado_inventario': estado_obj,
                        'ubicacion_actual': bodega,
                    }
                    if entregado_a_info:
                        defaults['entregado_a_info'] = entregado_a_info
                    if cliente_obj is not None:
                        defaults['cliente'] = cliente_obj
                        defaults['cliente_otro'] = ''
                    elif cliente_texto:
                        defaults['cliente_otro'] = cliente_texto
                    if correlativo not in (None, ''):
                        defaults['observaciones'] = f'Correlativo: {_as_text_id(correlativo)}'

                    medidor, created, reactivado = _con_reintento_sqlite(
                        lambda: _upsert_equipo(Medidor, 'serie', serie, defaults)
                    )
                    if created:
                        creadas += 1
                        registrar_movimiento_importacion(medidor, 'MEDIDOR', f'Serie {serie}', idx)
                    else:
                        actualizadas += 1
                    if reactivado:
                        avisos_reactivacion.append(
                            f'Fila {idx}: Medidor serie {serie} reactivado (estaba eliminado).'
                        )

                elif tipo == 'SIM':
                    imei = _as_text_id(get_val(valores, 'imei', 'icc', 'iccid'))
                    operador = _as_text_id(get_val(valores, 'operador'))
                    abonado = _as_text_id(get_val(valores, 'abonado', 'msisdn', 'telefono'))
                    apn = _as_text_id(get_val(valores, 'apn'))
                    fecha_recepcion_raw = get_val(valores, 'fecha_de_recepcion', 'fecha_recepcion')

                    if not imei:
                        raise ValueError(error_columna('IMEI', 'valor obligatorio'))
                    if not operador:
                        raise ValueError(error_columna('OPERADOR', 'valor obligatorio'))
                    # ABONADO puede venir vacío en maestros parciales; IMEI es la llave.
                    if not abonado:
                        abonado = ''
                    if not apn:
                        apn = ''
                    if not fecha_recepcion_raw:
                        # Permitir sin fecha: usar None
                        fecha_recepcion = None
                    else:
                        fecha_recepcion = _parse_fecha_flexible(fecha_recepcion_raw, 'FECHA RECEPCIÓN')
                    fecha_entrega = _parse_fecha_flexible(
                        get_val(valores, 'fecha_entrega', 'fecha_de_entrega'),
                        'FECHA ENTREGA',
                    )
                    estado_obj = _resolver_estado_inventario(get_val(valores, 'estado'), 'Instalado')
                    cliente_raw = get_val(valores, 'cliente')
                    cliente_obj = obtener_o_crear_cliente(cliente_raw, cache_clientes)
                    cliente_texto = _as_text_id(cliente_raw)
                    if cliente_texto.upper() in {'#N/A', 'N/A', 'NONE', 'NULL'}:
                        cliente_texto = ''
                    medidor_raw = get_val(valores, 'medidor')
                    medidor_obj = buscar_medidor(medidor_raw)
                    medidor_texto = _as_text_id(medidor_raw) if not medidor_obj else ''
                    entregado_a_nombre = _as_text_id(get_val(valores, 'entregado_a', 'entregado_a_'))
                    direccion_ip = _normalizar_direccion_ip(get_val(valores, 'direccion_ip', 'ip'))

                    defaults = {
                        'operador': operador,
                        'abonado': abonado,
                        'direccion_ip': direccion_ip,
                        'apn': apn,
                        'fecha_recepcion': fecha_recepcion,
                        'fecha_entrega': fecha_entrega,
                        'estado_inventario': estado_obj,
                        'ubicacion_actual': bodega,
                    }
                    # No borrar "Entregado a" de MoreApp/manual si el Excel viene vacío.
                    if entregado_a_nombre:
                        defaults['entregado_a_nombre'] = entregado_a_nombre
                    if cliente_obj is not None:
                        defaults['cliente'] = cliente_obj
                        defaults['cliente_otro'] = ''
                    elif cliente_texto:
                        defaults['cliente_otro'] = cliente_texto
                    if medidor_obj is not None:
                        defaults['medidor'] = medidor_obj
                        defaults['medidor_otro'] = ''
                    elif medidor_texto:
                        defaults['medidor_otro'] = medidor_texto
                    sim, created, reactivado = _con_reintento_sqlite(
                        lambda: _upsert_equipo(SimCard, 'imei', imei, defaults)
                    )
                    if created:
                        creadas += 1
                        registrar_movimiento_importacion(sim, 'SIM', f'IMEI {imei}', idx)
                    else:
                        actualizadas += 1
                    if reactivado:
                        avisos_reactivacion.append(
                            f'Fila {idx}: SIM IMEI {imei} reactivada (estaba eliminada).'
                        )

                elif tipo == 'MODEMS':
                    marca = _as_text_id(get_val(valores, 'marca'))
                    # Hay dos columnas "Marca" en el maestro; la primera es fabricante del módem.
                    # Si idx encuentra la secundaria, preferimos la primera aparición real.
                    if headers_norm.count('marca') > 1:
                        i0 = headers_norm.index('marca')
                        if i0 < len(valores):
                            marca = _as_text_id(valores[i0]) or marca

                    modelo = _as_text_id(get_val(valores, 'modelo'))
                    imei = _as_text_id(get_val(valores, 'imei')) or None
                    serie = _as_text_id(get_val(valores, 'serie'))
                    # Primera columna Serie (no la secundaria "Serie" de retirado)
                    if headers_norm.count('serie') > 1:
                        i0 = headers_norm.index('serie')
                        if i0 < len(valores):
                            serie = _as_text_id(valores[i0]) or serie

                    if not marca:
                        # Filas vacías / basura del maestro
                        if not serie and not imei:
                            contador_filas -= 1
                            continue
                        marca = 'SIN MARCA'
                    if not serie:
                        raise ValueError(error_columna('SERIE', 'valor obligatorio'))

                    fecha_recepcion = _parse_fecha_flexible(
                        get_val(valores, 'fecha_de_recepcion', 'fecha_recepcion'),
                        'Fecha de Recepción',
                    )
                    fecha_entrega = _parse_fecha_flexible(
                        get_val(valores, 'fecha_entrega', 'fecha_de_entrega'),
                        'Fecha Entrega',
                    )
                    caja_raw = _normalizar_caja_modem(get_val(valores, 'caja'))
                    if caja_raw:
                        caja_modem_actual = caja_raw
                    caja = caja_modem_actual or None
                    tecnico = _as_text_id(get_val(valores, 'tecnico_responsable', 'tecnico'))
                    cliente_raw = get_val(valores, 'cliente')
                    cliente_obj = obtener_o_crear_cliente(cliente_raw, cache_clientes)
                    cliente_texto = _as_text_id(cliente_raw)
                    if cliente_texto.upper() in {'#N/A', 'N/A', 'NONE', 'NULL'}:
                        cliente_texto = ''
                    medidor_raw = get_val(valores, 'medidor')
                    medidor_obj = buscar_medidor(medidor_raw)
                    medidor_texto = _as_text_id(medidor_raw) if not medidor_obj else ''

                    ip = _normalizar_direccion_ip(get_val(valores, 'ip', 'direccion_ip'))
                    puerto = _as_text_id(get_val(valores, 'puerto'))
                    # Marca secundaria = segunda columna "Marca" si existe
                    marca_secundaria = ''
                    if headers_norm.count('marca') > 1:
                        i1 = [i for i, h in enumerate(headers_norm) if h == 'marca'][1]
                        if i1 < len(valores):
                            marca_secundaria = _as_text_id(valores[i1])

                    # En el maestro la columna se llama "Estado" (antes Obs): es texto libre
                    # que se muestra como Obs en la página. Status suele venir vacío.
                    texto_estado_obs = _as_text_id(
                        get_val(valores, 'obs', 'observaciones', 'observacion', 'estado')
                    )
                    status_col = _as_text_id(get_val(valores, 'status'))
                    observaciones = texto_estado_obs
                    retirado = _as_text_id(get_val(valores, 'retirado'))
                    serie_secundaria = ''
                    if headers_norm.count('serie') > 1:
                        i1 = [i for i, h in enumerate(headers_norm) if h == 'serie'][1]
                        if i1 < len(valores):
                            serie_secundaria = _as_text_id(valores[i1])
                    irregularidad = _as_text_id(get_val(valores, 'irregularidad'))
                    proyecto = _as_text_id(get_val(valores, 'proyecto'))

                    # Estado de inventario: Status si existe; si no, se infiere del texto Estado/Obs
                    estado_raw = status_col or texto_estado_obs
                    estado_obj = _resolver_estado_inventario(estado_raw, 'En bodega')

                    if imei:
                        # Duplicado dentro del mismo Excel: el primero conserva el IMEI.
                        otra_serie = imeis_modem_en_import.get(imei)
                        if otra_serie and otra_serie != serie:
                            observaciones = (
                                f'{observaciones} | IMEI duplicado en Excel (también serie {otra_serie})'
                            ).strip(' |')
                            imei = None
                        else:
                            # Si el IMEI estaba en otro módem de BD, se libera (la serie del Excel manda).
                            Modem.objects.filter(imei=imei).exclude(serie=serie).update(imei=None)
                            imeis_modem_en_import[imei] = serie

                    defaults = {
                        'marca': marca,
                        'modelo': modelo,
                        'imei': imei,
                        'fecha_recepcion': fecha_recepcion,
                        'fecha_entrega': fecha_entrega,
                        'caja': caja,
                        'tecnico_responsable': tecnico,
                        'observaciones': observaciones,
                        'ip': ip,
                        'puerto': puerto,
                        'marca_secundaria': marca_secundaria,
                        'retirado': retirado,
                        'serie_secundaria': serie_secundaria,
                        'irregularidad': irregularidad,
                        'proyecto': proyecto,
                        'estado_inventario': estado_obj,
                        'ubicacion_actual': bodega,
                    }
                    if cliente_obj is not None:
                        defaults['cliente'] = cliente_obj
                        defaults['cliente_otro'] = ''
                    elif cliente_texto:
                        defaults['cliente_otro'] = cliente_texto
                    if medidor_obj is not None:
                        defaults['medidor'] = medidor_obj
                        defaults['medidor_otro'] = ''
                    elif medidor_texto:
                        defaults['medidor_otro'] = medidor_texto
                    modem, created, reactivado = _con_reintento_sqlite(
                        lambda: _upsert_equipo(Modem, 'serie', serie, defaults)
                    )
                    if created:
                        creadas += 1
                        registrar_movimiento_importacion(modem, 'MODEM', f'Serie {serie}', idx)
                    else:
                        actualizadas += 1
                    if reactivado:
                        avisos_reactivacion.append(
                            f'Fila {idx}: Módem serie {serie} reactivado (estaba eliminado).'
                        )
                else:
                    raise ValueError(f'Tipo de equipo no soportado: {tipo_equipo}')

                exitosas += 1
            except (IntegrityError, OperationalError) as exc:
                fallidas += 1
                ImportacionExcelError.objects.create(
                    importacion=importacion,
                    numero_fila=idx,
                    motivo=str(exc),
                    data_cruda=fila_a_texto(headers, valores),
                )
            except Exception as exc:
                fallidas += 1
                ImportacionExcelError.objects.create(
                    importacion=importacion,
                    numero_fila=idx,
                    motivo=str(exc),
                    data_cruda=fila_a_texto(headers, valores),
                )

        importacion.total_filas = contador_filas
        importacion.exitosas = exitosas
        importacion.fallidas = fallidas
        importacion.estado = 'COMPLETADO'
        obs = (
            f'Hoja "{ws.title}": {exitosas}/{contador_filas} OK '
            f'(nuevos: {creadas}, actualizados: {actualizadas}, fallidos: {fallidas})'
        )
        if avisos_reactivacion:
            obs += f'. Reactivados desde soft-delete: {len(avisos_reactivacion)}'
        importacion.observaciones = obs
        importacion.save()
        importacion.warnings = list(avisos_reactivacion)

    except Exception as exc:
        importacion.estado = 'ERROR'
        importacion.observaciones = f'Error en importación: {str(exc)}'
        importacion.save()

    return importacion


def importar_clientes_excel(archivo, usuario, sincronizar_completo=False):
    """
    Importa clientes desde archivo Excel.

    Formato esperado:
    Columnas mínimas requeridas: Sector, Tipo Suministro, Numero Cliente, Comuna, Nombre Cliente,
    Dirección de Instalación, Marca Medidor, Proyecto, Serie Medidor.
    Columnas opcionales: Referencia, Ciudad, Ultimo Acceso, Ultimo Perfil Carga,
    Ultimo Perfil Instrumentacion, Ultimo Reset, Ultimo Registro Facturacion, Nota, Trabajo,
    IP, Puerto, Modem, Fecha Registro.
    """

    importacion = ImportacionExcel.objects.create(
        tipo='CLIENTES',
        archivo_original=archivo.name if hasattr(archivo, 'name') else 'Upload',
        usuario=usuario,
    )

    def error_columna(columna, mensaje, valor=None):
        base = f'Columna "{columna}": {mensaje}'
        if valor is not None and str(valor).strip() != '':
            return f'{base}. Valor recibido: {valor}'
        return base

    def fila_a_texto(headers_fila, valores_fila):
        data = {}
        for i, valor in enumerate(valores_fila):
            nombre_columna = headers_fila[i] if i < len(headers_fila) else None
            nombre_columna = str(nombre_columna).strip() if nombre_columna else f'Columna_{i + 1}'
            data[nombre_columna] = valor
        return json.dumps(data, ensure_ascii=False, default=str)

    def normalizar_header(valor):
        if valor is None:
            return ''
        texto = str(valor).strip().lower()
        texto = texto.replace('\t', ' ').replace('\r', ' ').replace('\n', ' ')
        texto = ' '.join(texto.split())
        texto = texto.replace('°', 'o').replace('º', 'o')
        texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('ascii')
        texto = texto.replace(':', '').replace('#', 'numero')
        texto = texto.replace('-', ' ').replace('/', ' ').replace('.', ' ').replace('_', ' ')
        texto = ' '.join(texto.split())
        texto = texto.replace(' ', '_')
        while '__' in texto:
            texto = texto.replace('__', '_')
        return texto

    def matches_header(header, alias):
        header_n = normalizar_header(header)
        alias_n = normalizar_header(alias)
        if alias_n == header_n:
            return True

        # Evitar falsos positivos con aliases cortos (ej. "ip" dentro de "tipo_suministro").
        if len(alias_n) <= 3 or len(header_n) <= 3:
            return False

        alias_parts = [p for p in alias_n.split('_') if p]
        header_parts = [p for p in header_n.split('_') if p]

        if alias_parts and header_parts and all(part in header_parts for part in alias_parts):
            return True

        return False

    def idx_col(*aliases):
        for alias in aliases:
            for idx, header in enumerate(headers_norm):
                if header is None:
                    continue
                if matches_header(header, alias):
                    return idx
        return None

    def get_val(*aliases):
        idx = idx_col(*aliases)
        if idx is None or idx >= len(valores):
            return None
        return valores[idx]

    def clean_text(valor):
        if valor is None:
            return None
        valor = str(valor).strip()
        return valor or None

    def normalizar_numero_cliente(valor):
        """Normaliza Numero Cliente para evitar variantes como 100.0 vs 100."""
        if valor is None:
            return None

        if isinstance(valor, (int,)):
            return str(valor)

        if isinstance(valor, float):
            if valor.is_integer():
                return str(int(valor))
            return str(valor).strip()

        texto = str(valor).strip()
        if not texto:
            return None

        # Excel suele exportar enteros como texto terminado en .0
        if texto.endswith('.0') and texto[:-2].isdigit():
            return texto[:-2]

        return texto

    def normalizar_clave(valor):
        """Normaliza textos para comparar duplicados sin distinguir mayúsculas/espacios."""
        if valor is None:
            return None
        texto = str(valor).strip().lower()
        return texto or None

    def parse_fecha_registro(valor):
        from datetime import datetime, date
        if isinstance(valor, date):
            return valor
        if valor is None:
            return None
        if isinstance(valor, str):
            valor = valor.strip()
            if not valor:
                return None
            formatos = ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y']
            for fmt in formatos:
                try:
                    return datetime.strptime(valor, fmt).date()
                except Exception:
                    continue
        if hasattr(valor, 'date'):
            try:
                return valor.date()
            except Exception:
                pass
        raise ValueError(error_columna('Fecha Registro', 'formato no válido. Usa DD/MM/AAAA o YYYY-MM-DD', valor))

    try:
        wb = openpyxl.load_workbook(archivo)

        def row_headers(fila):
            return [normalizar_header(cell.value) for cell in fila]

        def tokenize_header(valor):
            texto = normalizar_header(valor)
            tokens = [token for token in texto.split('_') if token]
            return set(tokens)

        def header_matches(alias, header):
            alias_tokens = tokenize_header(alias)
            header_tokens = tokenize_header(header)
            if not alias_tokens or not header_tokens:
                return False
            if alias_tokens == header_tokens:
                return True
            if alias_tokens.issubset(header_tokens) or header_tokens.issubset(alias_tokens):
                return True
            if alias_tokens & header_tokens:
                # require at least two shared meaningful tokens to avoid accidental matches
                return len(alias_tokens & header_tokens) >= 2
            return False

        def score_header_row(fila):
            """Asigna puntaje a una fila candidata de encabezados para elegir la mejor."""
            nombres = [cell.value for cell in fila]
            grupos = {
                'numero_cliente': ['numero_cliente', 'numero cliente', 'numero', 'client number', '#', 'nro_cliente', 'nro cliente', 'n° cliente', 'nº cliente', 'no_cliente', 'no cliente'],
                'sector': ['sector', 'setor'],
                'tipo_suministro': ['tipo_suministro', 'tipo suministro', 'tipo de suministro', 'suministro'],
                'comuna': ['comuna', 'municipio'],
                'customer_name': ['customer_name', 'nombre cliente', 'nombre del cliente', 'customer name', 'cliente', 'name'],
                'installation_address': ['installation_address', 'direccion instalacion', 'direccion de instalacion', 'direcion de instalacion', 'dirección instalación', 'dirección de instalación', 'installation address', 'direccion', 'dirección', 'direccion cliente'],
                'meter_manufacturer_id': ['meter_manufacturer_id', 'fabricante', 'fabricante/id medidor', 'fabricante id medidor', 'meter manufacturer id', 'marca medidor', 'marca del medidor', 'marca'],
                'proyecto': ['proyecto', 'project', 'nombre proyecto'],
                'meter_serial_n_1': ['meter_serial_n_1', 'serie medidor', 'serie del medidor', 'medidor serie', 'serial', 'serie'],
                'ip': ['ip', 'direccion ip', 'ip address', 'ipv4'],
            }

            matched_groups = 0
            has_numero_cliente = False

            for key, aliases in grupos.items():
                found = False
                for alias in aliases:
                    for header in nombres:
                        if header_matches(alias, header):
                            found = True
                            break
                    if found:
                        break

                if found:
                    matched_groups += 1
                    if key == 'numero_cliente':
                        has_numero_cliente = True

            if not has_numero_cliente:
                return 0

            return matched_groups

        def find_header_sheet():
            best_ws = None
            best_row = None
            best_score = 0

            for worksheet in wb.worksheets:
                for i in range(1, min(12, worksheet.max_row) + 1):
                    fila = worksheet[i]
                    score = score_header_row(fila)
                    if score > best_score:
                        best_score = score
                        best_ws = worksheet
                        best_row = i

            # Requiere al menos numero_cliente y algunas columnas adicionales
            if best_score >= 3:
                return best_ws, best_row
            return None, None

        ws, header_row_index = find_header_sheet()
        if ws is None:
            first_sheet = wb.worksheets[0]
            headers = [cell.value for cell in first_sheet[1]]
            raise ValueError(
                'No se encontraron encabezados válidos en el Excel. ' +
                f'Encabezados detectados en la primera hoja: {headers}'
            )

        headers = [cell.value for cell in ws[header_row_index]]
        headers_norm = [normalizar_header(h) for h in headers]

        columnas_relevantes = [
            idx_col('numero_cliente', 'numero cliente', 'numero de cliente', 'número de cliente', 'numero', 'nro_cliente', 'nro cliente'),
            idx_col('sector', 'setor'),
            idx_col('tipo_suministro', 'tipo suministro', 'tipo de suministro', 'suministro'),
            idx_col('comuna', 'municipio'),
            idx_col('customer_name', 'nombre cliente', 'nombre de cliente', 'nombre del cliente', 'customer name', 'cliente'),
            idx_col('installation_address', 'direccion instalacion', 'direccion de instalacion', 'direcion de instalacion', 'dirección instalación', 'dirección de instalación', 'installation address', 'direccion', 'dirección'),
            idx_col('meter_manufacturer_id', 'fabricante', 'marca medidor', 'marca del medidor', 'marca'),
            idx_col('proyecto', 'project', 'nombre proyecto'),
            idx_col('meter_serial_n_1', 'serie medidor', 'serie del medidor', 'medidor serie', 'serial', 'serie'),
            idx_col('ip', 'direccion ip', 'ip address', 'ipv4'),
        ]
        columnas_relevantes = [idx for idx in columnas_relevantes if idx is not None]

        ips_importadas = {}
        series_importadas = {}

        contador_filas = 0
        exitosas = 0
        fallidas = 0
        creadas = 0
        actualizadas = 0
        desactivadas = 0
        numeros_excel = set()
        numeros_en_filas = set()
        filas_numero_repetido = 0
        advertencias = []
        duplicados_ip_count = 0
        duplicados_serie_count = 0
        duplicados_numero_set = set()
        numero_cliente_frecuencia = {}
        numero_cliente_filas = {}
        numero_cliente_filas_detalle = {}
        proyecto_detectado_count = 0
        serie_detectada_count = 0

        for idx, row in enumerate(ws.iter_rows(min_row=header_row_index + 1, values_only=False), start=header_row_index + 1):

            try:
                valores = [cell.value for cell in row]

                def valor_no_vacio(v):
                    if v is None:
                        return False
                    return str(v).strip() != ''

                hay_dato_relevante = any(
                    col < len(valores) and valor_no_vacio(valores[col])
                    for col in columnas_relevantes
                )
                if not hay_dato_relevante:
                    continue

                contador_filas += 1

                numero = clean_text(get_val(
                    'numero_cliente', 'numero cliente', 'numero', 'client number', '#',
                    'nro_cliente', 'nro cliente', 'numero_cliente', 'numero cliente', 'n° cliente', 'nº cliente', 'no_cliente', 'no cliente',
                    'numero de cliente', 'número de cliente'
                ))
                sector = clean_text(get_val('sector', 'setor', 'sector facturacion', 'sector_facturacion'))
                tipo_suministro = clean_text(get_val('tipo_suministro', 'tipo suministro', 'tipo de suministro', 'suministro'))
                comuna = clean_text(get_val('comuna', 'municipio'))
                customer_name = clean_text(get_val('customer_name', 'nombre cliente', 'nombre de cliente', 'nombre del cliente', 'customer name', 'cliente'))
                installation_address = clean_text(get_val(
                    'installation_address', 'direccion instalacion', 'direccion de instalacion', 'direcion de instalacion',
                    'dirección instalación', 'dirección de instalación', 'installation address', 'direccion', 'dirección', 'direccion cliente'
                ))
                meter_manufacturer_id = clean_text(get_val(
                    'meter_manufacturer_id', 'fabricante', 'fabricante/id medidor', 'fabricante id medidor',
                    'meter manufacturer id', 'marca medidor', 'marca del medidor', 'marca'
                ))
                proyecto = clean_text(get_val(
                    'proyecto', 'project', 'nombre proyecto', 'proyecto cliente', 'proy', 'proyecto_id', 'id proyecto'
                ))
                meter_serial_n_1 = clean_text(get_val(
                    'meter_serial_n_1', 'serie medidor', 'serie del medidor', 'medidor serie',
                    'serial', 'serie', 'n serie', 'numero serie', 'nro serie', 'no serie',
                    'n serie medidor', 'numero serie medidor', 'nro serie medidor', 'serial medidor',
                    'numero medidor', 'nro medidor', 'medidor numero'
                ))
                referencia = clean_text(get_val('referencia', 'observaciones', 'notes', 'nota'))
                city = clean_text(get_val('city', 'ciudad'))
                ultimo_acceso = clean_text(get_val('ultimo_acceso', 'ultimo acceso'))
                ultimo_perfil_carga = clean_text(get_val('ultimo_perfil_carga', 'ultimo perfil carga'))
                ultimo_perfil_instrumentacion = clean_text(get_val('ultimo_perfil_instrumentacion', 'ultimo perfil instrumentacion'))
                ultimo_reset = clean_text(get_val('ultimo_reset', 'ultimo reset'))
                ultimo_registro_facturacion = clean_text(get_val('ultimo_registro_facturacion', 'ultimo registro facturacion'))
                note = clean_text(get_val('note', 'nota', 'notas'))
                trabajo = clean_text(get_val('trabajo', 'work'))
                client_type = clean_text(get_val('client_type', 'client type', 'tipo cliente', 'tipo de cliente'))
                ip_raw = get_val('ip', 'direccion ip', 'ip address', 'ipv4')
                ip = normalize_ip_value(ip_raw)
                if ip_raw is not None and str(ip_raw).strip() != '' and ip and str(ip_raw).strip() != str(ip):
                    advertencias.append(
                        f'Fila {idx}: [IP] se normalizó de "{ip_raw}" a "{ip}" (formato Excel sin puntos).'
                    )
                puerto = clean_text(get_val('puerto', 'port'))
                modem = clean_text(get_val('modem'))
                fecha_registro_raw = clean_text(get_val('fecha_registro', 'fecha de registro', 'registro fecha'))

                validation_issues = merge_issues(
                    validate_ip_format(ip),
                    validate_ip_port_coherence(ip, puerto),
                )
                for issue in validation_issues:
                    if issue.severity == 'error' and issue.code == 'IP_INVALID_FORMAT':
                        # No tumbar la fila: el cliente importa; IP queda vacía con aviso
                        advertencias.append(
                            f'Fila {idx}: [VALIDACION] {issue.message}. Se importó sin IP.'
                        )
                        ip = None
                    elif issue.severity == 'error':
                        raise ValueError(error_columna('IP', issue.message, ip))
                    else:
                        advertencias.append(f'Fila {idx}: [VALIDACION] {issue.message}')

                if not numero:
                    raise ValueError('Falta campo obligatorio: Numero Cliente')

                numero_text = normalizar_numero_cliente(numero)
                if not numero_text:
                    raise ValueError('Falta campo obligatorio: Numero Cliente')

                if numero_text in numeros_en_filas:
                    filas_numero_repetido += 1
                    duplicados_numero_set.add(numero_text)
                else:
                    numeros_en_filas.add(numero_text)

                numero_cliente_frecuencia[numero_text] = numero_cliente_frecuencia.get(numero_text, 0) + 1
                numero_cliente_filas.setdefault(numero_text, []).append(idx)
                numero_cliente_filas_detalle.setdefault(numero_text, []).append({
                    'fila_excel': idx,
                    'sector': sector or '',
                    'tipo_suministro': tipo_suministro or '',
                    'numero_cliente': numero_text,
                    'comuna': comuna or '',
                    'customer_name': customer_name or '',
                    'installation_address': installation_address or '',
                    'meter_manufacturer_id': meter_manufacturer_id or '',
                    'proyecto': (proyecto or 'SIN PROYECTO') or '',
                    'meter_serial_n_1': meter_serial_n_1 or '',
                })

                if numero_text == '0':
                    raise ValueError(error_columna('Numero Cliente', 'valor inválido', numero_text))

                if sector and str(sector).strip() == '0':
                    sector = None
                serie_text = _normalizar_serie_medidor_cliente(meter_serial_n_1)
                if meter_serial_n_1 and serie_text != str(meter_serial_n_1).strip():
                    meter_serial_n_1 = serie_text
                elif serie_text:
                    meter_serial_n_1 = serie_text
                else:
                    meter_serial_n_1 = None

                cliente_existente, reactivar = _buscar_cliente_para_importacion(numero_text, serie_text)
                if cliente_existente and reactivar:
                    cliente_existente.activo = True
                    cliente_existente.fecha_eliminacion = None
                    cliente_existente.eliminado_por = None
                    advertencias.append(
                        f'Fila {idx}: Cliente {numero_text} reactivado (estaba inactivo/eliminado).'
                    )
                elif cliente_existente and serie_text and not _serie_cliente_igual(
                    cliente_existente.meter_serial_n_1, serie_text
                ):
                    advertencias.append(
                        f'Fila {idx}: Cliente {numero_text} actualizado '
                        f'(serie {cliente_existente.meter_serial_n_1 or "—"} → {serie_text}).'
                    )

                ip_key = normalizar_clave(ip)
                if ip_key:
                    if Cliente.objects.filter(activo=True, ip__iexact=ip).exclude(numero_cliente=numero_text).exists():
                        duplicados_ip_count += 1
                        advertencias.append(
                            f'Fila {idx}: ' + error_columna('IP', '[DUPLICADO] ya existe en clientes activos', ip)
                        )
                    if ip_key in ips_importadas:
                        fila_prev, numero_prev = ips_importadas[ip_key]
                        if numero_prev != numero_text:
                            duplicados_ip_count += 1
                            advertencias.append(
                                f'Fila {idx}: ' + error_columna('IP', f'[DUPLICADO] repetida en el mismo Excel (primera aparición en fila {fila_prev})', ip)
                            )

                serie_key = normalizar_clave(meter_serial_n_1)
                if serie_key:
                    meter_exists_other_active = Cliente.objects.filter(
                        activo=True,
                        meter_serial_n_1__iexact=meter_serial_n_1,
                    ).exclude(numero_cliente=numero_text).exists()
                    meter_issues = validate_meter_uniqueness(meter_serial_n_1, meter_exists_other_active)
                    for issue in meter_issues:
                        duplicados_serie_count += 1
                        advertencias.append(
                            f'Fila {idx}: [VALIDACION] {issue.message}'
                        )
                    if serie_key in series_importadas:
                        fila_prev, numero_prev = series_importadas[serie_key]
                        if numero_prev != numero_text:
                            duplicados_serie_count += 1
                            advertencias.append(
                                f'Fila {idx}: ' + error_columna('Serie Medidor', f'[DUPLICADO] repetida en el mismo Excel (primera aparición en fila {fila_prev})', meter_serial_n_1)
                            )

                medidor_obj = None
                if meter_serial_n_1:
                    serie_detectada_count += 1
                    medidor_text = str(meter_serial_n_1).strip()
                    medidor_obj = Medidor.objects.filter(serie__iexact=medidor_text, eliminado=False).first()
                    # La serie se guarda en Cliente aunque el medidor no exista en inventario.
                    # Solo validamos asignación duplicada cuando sí existe objeto Medidor.
                    if medidor_obj and Cliente.objects.filter(medidor_actual=medidor_obj, activo=True).exclude(numero_cliente=numero_text).exists():
                        duplicados_serie_count += 1
                        advertencias.append(
                            f'Fila {idx}: ' + error_columna('Serie Medidor', '[DUPLICADO] el medidor está asignado a otro cliente activo', meter_serial_n_1)
                        )
                        medidor_obj = None

                modem_issues = validate_modem_assignment(
                    modem,
                    bool(modem and Cliente.objects.filter(activo=True, modem__iexact=modem).exclude(numero_cliente=numero_text).exists()),
                )
                for issue in modem_issues:
                    advertencias.append(f'Fila {idx}: [VALIDACION] {issue.message}')

                fecha_registro = parse_fecha_registro(fecha_registro_raw)

                if proyecto:
                    proyecto_detectado_count += 1
                proyecto_final = proyecto or 'SIN PROYECTO'

                if cliente_existente:
                    if sector:
                        cliente_existente.sector = sector
                    if tipo_suministro:
                        cliente_existente.tipo_suministro = tipo_suministro
                    if comuna:
                        cliente_existente.comuna = comuna
                    if customer_name:
                        cliente_existente.customer_name = customer_name
                    if installation_address:
                        cliente_existente.installation_address = installation_address
                        cliente_existente.direccion = installation_address
                    if meter_manufacturer_id:
                        cliente_existente.meter_manufacturer_id = meter_manufacturer_id
                    if proyecto:
                        cliente_existente.proyecto = proyecto
                    elif not cliente_existente.proyecto:
                        cliente_existente.proyecto = proyecto_final
                    if meter_serial_n_1:
                        cliente_existente.meter_serial_n_1 = meter_serial_n_1
                    if referencia:
                        cliente_existente.referencia = referencia
                    if city:
                        cliente_existente.city = city
                    if client_type:
                        cliente_existente.client_type = client_type
                    if ultimo_acceso:
                        cliente_existente.ultimo_acceso = ultimo_acceso
                    if ultimo_perfil_carga:
                        cliente_existente.ultimo_perfil_carga = ultimo_perfil_carga
                    if ultimo_perfil_instrumentacion:
                        cliente_existente.ultimo_perfil_instrumentacion = ultimo_perfil_instrumentacion
                    if ultimo_reset:
                        cliente_existente.ultimo_reset = ultimo_reset
                    if ultimo_registro_facturacion:
                        cliente_existente.ultimo_registro_facturacion = ultimo_registro_facturacion
                    if note:
                        cliente_existente.note = note
                    if trabajo:
                        cliente_existente.trabajo = trabajo
                    if ip:
                        cliente_existente.ip = ip
                    if puerto:
                        cliente_existente.puerto = puerto
                    if modem:
                        cliente_existente.modem = modem
                    if fecha_registro_raw:
                        cliente_existente.fecha_registro = fecha_registro
                    if medidor_obj:
                        cliente_existente.medidor_actual = medidor_obj
                    cliente_existente.save()
                    actualizadas += 1
                else:
                    Cliente.objects.create(
                        numero_cliente=numero_text,
                        direccion=installation_address or '',
                        comuna=comuna or '',
                        referencia=referencia,
                        tipo_suministro=tipo_suministro,
                        sector=sector,
                        city=city,
                        customer_name=customer_name,
                        installation_address=installation_address or '',
                        proyecto=proyecto_final,
                        meter_manufacturer_id=meter_manufacturer_id,
                        meter_serial_n_1=meter_serial_n_1,
                        client_type=client_type,
                        ultimo_acceso=ultimo_acceso,
                        ultimo_perfil_carga=ultimo_perfil_carga,
                        ultimo_perfil_instrumentacion=ultimo_perfil_instrumentacion,
                        ultimo_reset=ultimo_reset,
                        ultimo_registro_facturacion=ultimo_registro_facturacion,
                        note=note,
                        trabajo=trabajo,
                        ip=ip,
                        puerto=puerto,
                        modem=modem,
                        fecha_registro=fecha_registro,
                        medidor_actual=medidor_obj,
                        activo=True,
                    )
                    creadas += 1
                if ip_key:
                    ips_importadas[ip_key] = (idx, numero_text)
                if serie_key:
                    series_importadas[serie_key] = (idx, numero_text)

                # Clave post-save: si el Excel trae serie vacía pero la ficha ya tenía
                # serie, no usar '' (eso desactivaría el cliente en sync completo).
                if cliente_existente:
                    serie_sync = cliente_existente.meter_serial_n_1
                else:
                    serie_sync = meter_serial_n_1
                numeros_excel.add(_clave_cliente_sync(numero_text, serie_sync))

                exitosas += 1

            except Exception as e:
                fallidas += 1
                ImportacionExcelError.objects.create(
                    importacion=importacion,
                    numero_fila=idx,
                    motivo=str(e),
                    data_cruda=fila_a_texto(headers, valores)
                )

        desactivadas = 0
        if sincronizar_completo and numeros_excel:
            for cliente_activo in Cliente.objects.filter(activo=True):
                clave_cliente = _clave_cliente_sync(
                    cliente_activo.numero_cliente,
                    cliente_activo.meter_serial_n_1,
                )
                if clave_cliente not in numeros_excel:
                    cliente_activo.activo = False
                    cliente_activo.save(update_fields=['activo'])
                    desactivadas += 1

        importacion.total_filas = contador_filas
        importacion.exitosas = exitosas
        importacion.fallidas = fallidas
        importacion.estado = 'COMPLETADO' if exitosas > 0 else 'ERROR'
        resumen_modo = 'sincronización completa' if sincronizar_completo else 'importación incremental'
        importacion.observaciones = (
            f'Importación finalizada. Filas útiles: {contador_filas}. '
            f'Correctas: {exitosas}. Con error: {fallidas}. '
            f'Clientes únicos detectados en archivo: {len(numeros_en_filas)}. '
            f'Filas repetidas por Numero Cliente: {filas_numero_repetido}. '
            f'Proyecto detectado en {proyecto_detectado_count} filas. '
            f'Serie Medidor detectada en {serie_detectada_count} filas. '
            f'Resultado de {resumen_modo}: {creadas} creados, {actualizadas} actualizados, {desactivadas} desactivados. '
            f'Advertencias de duplicado: {len(advertencias)}.'
        )
        importacion.save()
        importacion.warnings = advertencias
        importacion.warning_summary = {
            'duplicados_total': len(advertencias),
            'duplicados_numero': len(duplicados_numero_set),
            'duplicados_ip': duplicados_ip_count,
            'duplicados_serie': duplicados_serie_count,
            'filas_numero_repetido': filas_numero_repetido,
            'clientes_unicos_detectados': len(numeros_en_filas),
            'numeros_duplicados_muestra': sorted(list(duplicados_numero_set))[:500],
            'duplicados_numero_detalle': [
                {
                    'numero': numero,
                    'repeticiones': frecuencia,
                    'filas': numero_cliente_filas.get(numero, []),
                }
                for numero, frecuencia in sorted(numero_cliente_frecuencia.items(), key=lambda x: (-x[1], x[0]))
                if frecuencia > 1
            ][:1000],
            'duplicados_numero_filas': {
                numero: numero_cliente_filas_detalle.get(numero, [])
                for numero, frecuencia in sorted(numero_cliente_frecuencia.items(), key=lambda x: (-x[1], x[0]))
                if frecuencia > 1
            },
        }
        register_audit_event(
            AuditEvent(
                actor_id=getattr(usuario, 'id', None),
                action='CLIENT_IMPORT',
                entity='ImportacionExcel',
                entity_id=str(importacion.id),
                field_name='estado',
                old_value=None,
                new_value=importacion.estado,
                reason=(
                    f'Importación clientes ({resumen_modo}) - '
                    f'creados={creadas}, actualizados={actualizadas}, '
                    f'desactivados={desactivadas}, errores={fallidas}'
                ),
            )
        )

    except Exception as e:
        importacion.estado = 'ERROR'
        importacion.observaciones = f'Error: {str(e)}'
        importacion.save()

    return importacion


def exportar_clientes_excel(clientes):
    """
    Exporta clientes a archivo Excel.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CLIENTES'

    headers = [
        'Sector',
        'Tipo Suministro',
        'Numero Cliente',
        'Comuna',
        'Nombre Cliente',
        'Dirección Instalación',
        'Marca Medidor',
        'Proyecto',
        'Serie Medidor',
    ]

    ws.append(headers)

    for cliente in clientes:
        row = [
            cliente.sector or '',
            cliente.tipo_suministro or '',
            cliente.numero_cliente,
            cliente.comuna,
            cliente.customer_name or '',
            cliente.installation_address or '',
            cliente.meter_manufacturer_id or '',
            cliente.proyecto or '',
            cliente.meter_serial_n_1 or '',
        ]
        ws.append(row)

    # Filtro solo en Sector / Tipo / Nº cliente / Comuna.
    # Sin filtro en Nombre, Dirección, Marca, Proyecto, Serie (texto largo o casi único).
    aplicar_estilo_hoja_exportacion(ws, auto_filter=True, filter_from_col=1, filter_to_col=4)
    return wb


def _serie_medidor_display(equipo) -> str:
    """Serie de medidor vinculada o texto importado (medidor_otro)."""
    medidor = getattr(equipo, 'medidor', None)
    serie = getattr(medidor, 'serie', None) if medidor else None
    if serie:
        return str(serie).strip()
    return str(getattr(equipo, 'medidor_otro', '') or '').strip()


def exportar_equipos_excel(equipos, tipo_equipo='MEDIDORES'):
    """
    Exporta equipos a archivo Excel.
    
    Args:
        equipos: QuerySet de equipos (Medidor, SimCard o Modem)
        tipo_equipo: Tipo de equipo ('MEDIDORES', 'SIM', 'MODEMS')
    
    Returns:
        openpyxl Workbook object
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo_equipo
    
    # Encabezados según tipo (mismos nombres/orden que usa la importación)
    if tipo_equipo.upper() == 'MEDIDORES':
        headers = [
            '#', 'FECHA DE RECEPCION', 'BODEGA', 'MARCA', 'CAJA', 'MEDIDOR', 'MODULO',
            'FECHA DE ENTREGA', 'ENTREGADO A:', 'ESTADO', 'CLIENTE', 'Tipo Medidor',
        ]
    elif tipo_equipo.upper() == 'SIM':
        headers = ['IMEI', 'OPERADOR', 'ABONADO', 'DIRECCIÓN IP', 'APN', 'FECHA RECEPCIÓN', 'ENTREGADO A', 'FECHA ENTREGA', 'ESTADO', 'CLIENTE', 'MEDIDOR']
    elif tipo_equipo.upper() == 'MODEMS':
        headers = ['MARCA', 'MODELO', 'IMEI', 'SERIE', 'Fecha Recepción', 'Fecha Entrega', 'Caja', 'Técnico Responsable', 'Cliente', 'Medidor', 'IP', 'Puerto', 'Marca Secundaria', 'Observaciones', 'Retirado', 'Serie Secundaria', 'Irregularidad', 'Proyecto']
    else:
        headers = ['Datos']
    
    ws.append(headers)
    
    for indice, equipo in enumerate(equipos, start=1):
        if tipo_equipo.upper() == 'MEDIDORES':
            row = [
                indice,
                equipo.fecha_recepcion.strftime('%d-%m-%Y') if equipo.fecha_recepcion else '',
                equipo.bodega or '',
                equipo.marca or '',
                equipo.caja or '',
                equipo.serie,
                'SI' if getattr(equipo, 'modulo', None) is True else ('NO' if getattr(equipo, 'modulo', None) is False else ''),
                equipo.fecha_entrega.strftime('%d-%m-%Y') if equipo.fecha_entrega else '',
                (
                    equipo.entregado_a.nombre_interno if getattr(equipo, 'entregado_a', None) else ''
                ) or (equipo.entregado_a_info or '') or (getattr(equipo, 'entregado_a_otro', '') or ''),
                equipo.estado_inventario.nombre if equipo.estado_inventario else '',
                (
                    equipo.cliente.numero_cliente if getattr(equipo, 'cliente', None) else ''
                ) or (getattr(equipo, 'cliente_otro', '') or ''),
                getattr(equipo, 'tipo_medidor', '') or '',
            ]
        elif tipo_equipo.upper() == 'SIM':
            row = [
                equipo.imei or '',
                equipo.operador or '',
                equipo.abonado or '',
                equipo.direccion_ip or '',
                equipo.apn or '',
                equipo.fecha_recepcion.strftime('%d-%m-%Y') if equipo.fecha_recepcion else '',
                (
                    getattr(equipo, 'entregado_a_nombre', '') or ''
                ) or (getattr(equipo, 'entregado_a_otro', '') or '') or (
                    equipo.en_custodia_de.nombre_interno if getattr(equipo, 'en_custodia_de', None) else ''
                ),
                equipo.fecha_entrega.strftime('%d-%m-%Y') if equipo.fecha_entrega else '',
                equipo.estado_inventario.nombre if equipo.estado_inventario else '',
                equipo.cliente.numero_cliente if equipo.cliente else (getattr(equipo, 'cliente_otro', '') or ''),
                _serie_medidor_display(equipo),
            ]
        elif tipo_equipo.upper() == 'MODEMS':
            row = [
                equipo.marca or '',
                equipo.modelo or '',
                equipo.imei or '',
                equipo.serie,
                equipo.fecha_recepcion.strftime('%d-%m-%Y') if equipo.fecha_recepcion else '',
                equipo.fecha_entrega.strftime('%d-%m-%Y') if equipo.fecha_entrega else '',
                equipo.caja or '',
                equipo.tecnico_responsable or '',
                equipo.cliente.numero_cliente if equipo.cliente else (getattr(equipo, 'cliente_otro', '') or ''),
                _serie_medidor_display(equipo),
                equipo.ip or '',
                equipo.puerto or '',
                equipo.marca_secundaria or '',
                equipo.observaciones or '',
                equipo.retirado or '',
                equipo.serie_secundaria or '',
                equipo.irregularidad or '',
                equipo.proyecto or ''
            ]
        else:
            row = ['']
        
        ws.append(row)

    # En medidores la columna '#' no aporta filtro; el resto sí.
    if tipo_equipo.upper() == 'MEDIDORES':
        aplicar_estilo_hoja_exportacion(ws, auto_filter=True, filter_from_col=2)
    else:
        aplicar_estilo_hoja_exportacion(ws, auto_filter=True)
    return wb


def _normalizar_valor_export_completo(valor):
    from datetime import date, datetime
    from decimal import Decimal

    if valor is None:
        return ''
    if isinstance(valor, bool):
        return 'SI' if valor else 'NO'
    if isinstance(valor, datetime):
        return valor.strftime('%d-%m-%Y %H:%M:%S')
    if isinstance(valor, date):
        return valor.strftime('%d-%m-%Y')
    if isinstance(valor, Decimal):
        return float(valor)
    return str(valor)


def _etiqueta_relacion_export_completo(campo, relacionado, obj=None):
    if relacionado is None:
        if campo.name == 'medidor' and obj is not None:
            return getattr(obj, 'medidor_otro', '') or ''
        return ''

    if campo.name == 'cliente':
        return getattr(relacionado, 'numero_cliente', '') or str(relacionado)
    if campo.name in ('entregado_a', 'en_custodia_de', 'eliminado_por'):
        return getattr(relacionado, 'nombre_interno', '') or getattr(relacionado, 'username', '') or str(relacionado)
    if campo.name == 'estado_inventario':
        return getattr(relacionado, 'nombre', '') or str(relacionado)
    if campo.name == 'ubicacion_actual':
        return str(relacionado)
    if campo.name in ('medidor', 'medidor_actual'):
        return getattr(relacionado, 'serie', '') or str(relacionado)
    return str(relacionado)


def llenar_hoja_modelo_completo(ws, objetos, model, excluir_campos=None, incluir_ids_relacion=True):
    """Vuelca en la hoja todos los campos concretos del modelo.

    Relaciones → columnas `_id` + `_label` (o solo `_label` si incluir_ids_relacion=False);
    choices → valor + `_display`.
    """
    normalizar_valor = _normalizar_valor_export_completo

    excluir_campos = set(excluir_campos or ())
    columnas = []
    claves = []

    for campo in model._meta.concrete_fields:
        if campo.name in excluir_campos:
            continue
        if campo.is_relation:
            if incluir_ids_relacion:
                claves.append(f'{campo.name}_id')
                columnas.append((f'{campo.name}_id', lambda obj, nombre=campo.attname: normalizar_valor(getattr(obj, nombre, None))))
            claves.append(f'{campo.name}_label')
            columnas.append((
                f'{campo.name}_label',
                lambda obj, campo_rel=campo: normalizar_valor(
                    _etiqueta_relacion_export_completo(campo_rel, getattr(obj, campo_rel.name, None), obj)
                )
            ))
        else:
            claves.append(campo.name)
            if getattr(campo, 'choices', None):
                columnas.append((campo.name, lambda obj, nombre=campo.name: normalizar_valor(getattr(obj, nombre, None))))
                claves.append(f'{campo.name}_display')
                columnas.append((
                    f'{campo.name}_display',
                    lambda obj, nombre=campo.name: normalizar_valor(getattr(obj, f'get_{nombre}_display')()) if getattr(obj, nombre, None) not in (None, '') else ''
                ))
            else:
                columnas.append((campo.name, lambda obj, nombre=campo.name: normalizar_valor(getattr(obj, nombre, None))))

    ws.append([_etiqueta_columna_export_completo(clave) for clave in claves])

    for obj in objetos:
        ws.append([funcion(obj) for _, funcion in columnas])

    # Completo: sin filtro (demasiadas columnas técnicas / IDs no filtrables de forma útil)
    aplicar_estilo_hoja_exportacion(ws, header_fill_hex='7F6000', max_width=36, auto_filter=False)


def exportar_clientes_excel_completo(clientes):
    """Exporta todos los datos vigentes del cliente en una única hoja."""
    from clientes.models import Cliente

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CLIENTES COMPLETOS'
    llenar_hoja_modelo_completo(
        ws,
        clientes,
        Cliente,
        excluir_campos={'id'},
        incluir_ids_relacion=False,
    )
    return wb


def exportar_equipos_excel_completo(equipos, tipo_equipo='MEDIDORES'):
    """Exporta todos los campos reales del modelo y columnas legibles de apoyo."""
    model = getattr(equipos, 'model', None)
    if model is None:
        raise ValueError('El exportador completo requiere un QuerySet del modelo')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{tipo_equipo}_COMPLETO'
    llenar_hoja_modelo_completo(ws, equipos, model)
    return wb


def generar_plantilla_modem():
    """
    Genera archivo Excel plantilla para importar módems.
    VERDE: Campos del Excel (solo lectura)
    AMARILLO: Campos editables por administrativo
    NARANJA: Campos ocultos (solo admin/auditor)
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'MODEMS'
    
    # Definir fills
    verde_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Verde claro
    amarillo_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')  # Amarillo claro
    naranja_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')  # Naranja/rosa
    bold_font = Font(bold=True)
    
    # Encabezados
    # VERDE (columnas A-H): MARCA, MODELO, IMEI, SERIE, Fecha Recepción, Fecha Entrega, Caja, Técnico
    verde_headers = ['MARCA', 'MODELO', 'IMEI', 'SERIE', 'Fecha Recepción', 'Fecha Entrega', 'Caja', 'Técnico Responsable']
    
    # AMARILLO (columnas I-J): Cliente, Medidor
    amarillo_headers = ['Cliente', 'Medidor']
    
    # NARANJA (columnas K-R): IP, Puerto, Marca Sec, Observaciones, Retirado, Serie Sec, Irregularidad, Proyecto
    naranja_headers = ['IP', 'Puerto', 'Marca Secundaria', 'Observaciones', 'Retirado', 'Serie Secundaria', 'Irregularidad', 'Proyecto']
    
    col_idx = 1
    
    # Escribir encabezados verdes
    for header in verde_headers:
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = verde_fill
        cell.font = bold_font
        col_idx += 1
    
    # Escribir encabezados amarillos
    for header in amarillo_headers:
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = amarillo_fill
        cell.font = bold_font
        col_idx += 1
    
    # Escribir encabezados naranjas
    for header in naranja_headers:
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = naranja_fill
        cell.font = bold_font
        col_idx += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 15  # MARCA
    ws.column_dimensions['B'].width = 15  # MODELO
    ws.column_dimensions['C'].width = 20  # IMEI
    ws.column_dimensions['D'].width = 20  # SERIE
    ws.column_dimensions['E'].width = 15  # Fecha Recepción
    ws.column_dimensions['F'].width = 15  # Fecha Entrega
    ws.column_dimensions['G'].width = 12  # Caja
    ws.column_dimensions['H'].width = 20  # Técnico
    ws.column_dimensions['I'].width = 15  # Cliente
    ws.column_dimensions['J'].width = 15  # Medidor
    ws.column_dimensions['K'].width = 15  # IP
    ws.column_dimensions['L'].width = 10  # Puerto
    ws.column_dimensions['M'].width = 15  # Marca Sec
    ws.column_dimensions['N'].width = 25  # Observaciones
    ws.column_dimensions['O'].width = 12  # Retirado
    ws.column_dimensions['P'].width = 20  # Serie Sec
    ws.column_dimensions['Q'].width = 20  # Irregularidad
    ws.column_dimensions['R'].width = 15  # Proyecto
    
    # Hoja de instrucciones
    ws_instruc = wb.create_sheet('Instrucciones')
    ws_instruc['A1'] = 'INSTRUCCIONES PARA IMPORTAR MÓDEMS'
    ws_instruc['A1'].font = Font(bold=True, size=14)
    
    instrucciones = [
        '',
        '1. CAMPOS VERDES (Columnas A-H): Datos que vienen del Excel',
        '   - MARCA: Marca del módem (obligatorio)',
        '   - MODELO: Modelo del módem',
        '   - IMEI: Código IMEI único',
        '   - SERIE: Número de serie único (obligatorio)',
        '   - Fecha Recepción: Fecha de recepción del equipo (DD-MM-YYYY)',
        '   - Fecha Entrega: Fecha de entrega al técnico (DD-MM-YYYY)',
        '   - Caja: Número de caja',
        '   - Técnico Responsable: Nombre del técnico',
        '',
        '2. CAMPOS AMARILLOS (Columnas I-K): Editables por administrativo',
        '   - Cliente: Número de cliente',
        '   - Cliente: Número de cliente',
        '   - Medidor: Serie del medidor',
        '',
        '3. CAMPOS NARANJAS (Columnas K-R): Ocultos para administrativo',
        '   - Solo visibles para admin/auditor',
        '   - Incluyen: IP, Puerto, Marca Secundaria, Observaciones, Retirado, Serie Secundaria, Irregularidad, Proyecto',
        '',
        '4. Notas importantes:',
        '   - Los campos MARCA y SERIE son obligatorios',
        '   - La SERIE debe ser única en el sistema',
        '   - Las fechas deben estar en formato DD-MM-YYYY',
        '   - Los campos amarillos pueden dejarse vacíos y llenarse después en el sistema',
        '   - Los campos naranjas son opcionales y solo para uso administrativo interno',
        '',
        f'Plantilla generada el {datetime.now().strftime("%d-%m-%Y %H:%M")}'
    ]
    
    for idx, instruccion in enumerate(instrucciones, start=2):
        ws_instruc[f'A{idx}'] = instruccion
    
    ws_instruc.column_dimensions['A'].width = 100
    
    # Guardar
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_plantilla_simcard():
    """
    Genera plantilla Excel para importar SIM Cards
    
    CAMPOS AMARILLOS (obligatorios - vienen desde planilla):
    - IMEI, OPERADOR, ABONADO, DIRECCIÓN IP, APN, FECHA DE RECEPCIÓN, ENTREGADO A
    
    CAMPOS VERDES (opcionales - modifica administrativo después):
    - FECHA ENTREGA, ESTADO, CLIENTE, MEDIDOR
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SIM Cards"
    
    # Estilos
    amarillo = "FFFF00"
    verde = "00B050"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    amarillo_fill = PatternFill(start_color=amarillo, end_color=amarillo, fill_type="solid")
    verde_fill = PatternFill(start_color=verde, end_color=verde, fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados - según imagen del usuario
    headers = [
        ('IMEI', amarillo),
        ('OPERADOR', amarillo),
        ('ABONADO', amarillo),
        ('DIRECCIÓN IP', amarillo),
        ('APN', amarillo),
        ('FECHA DE RECEPCIÓN', amarillo),
        ('ENTREGADO A', amarillo),
        ('FECHA ENTREGA', verde),
        ('ESTADO', verde),
        ('CLIENTE', verde),
        ('MEDIDOR', verde)
    ]
    
    # Agregar encabezados
    for idx, (header, color) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.alignment = center_alignment
        cell.border = border
    
    # Filas de ejemplo (vacías)
    for row_num in range(2, 12):  # 10 filas de ejemplo
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            if col_num <= 7:  # Campos amarillos (columnas 1-7)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            else:  # Campos verdes (columnas 8-11)
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Ajustar ancho de columnas
    widths = [18, 15, 18, 18, 25, 20, 18, 18, 15, 15, 15]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    
    # Agregar hoja de instrucciones
    ws_instruc = wb.create_sheet("Instrucciones")
    
    instrucciones = [
        ("GUÍA DE IMPORTACIÓN SIM CARDS", ""),
        ("", ""),
        ("CAMPOS AMARILLOS (Obligatorios - Debe venir en la planilla):", ""),
        ("IMEI", "Identificador único del equipo/modem (ej: 895601000013094)"),
        ("OPERADOR", "Operador telefónico (ENTEL, CLARO, MOVISTAR, VTECH, etc.)"),
        ("ABONADO", "Número de abonado o identificador del operador"),
        ("DIRECCIÓN IP", "Dirección IP asignada por el operador"),
        ("APN", "Access Point Name configurado (ej: Eneldx.entel.cl)"),
        ("FECHA DE RECEPCIÓN", "Fecha de recepción en bodega (formato: DD-MM-YYYY o Excel date)"),
        ("ENTREGADO A", "Nombre de la persona a quien se entregó (ej: S. Suazo, C. Suazo)"),
        ("", ""),
        ("CAMPOS VERDES (Opcionales - Administrativo completa después):", ""),
        ("FECHA ENTREGA", "Fecha exacta de entrega (se completa después de importar)"),
        ("ESTADO", "Estado del inventario: Instalado, Bodega, Retirado, etc."),
        ("CLIENTE", "Cliente al que está asignada la SIM (se asigna después)"),
        ("MEDIDOR", "Número de medidor asociado (se asigna después)"),
        ("", ""),
        ("NOTAS IMPORTANTES:", ""),
        ("1.", "IMEI debe ser único (no puede haber duplicados)"),
        ("2.", "Los campos amarillos se cargan desde la planilla original"),
        ("3.", "Los campos verdes se modifican usando el botón 'Modificar' en el sistema"),
        ("4.", "Todas las SIM se crean con ESTADO='Instalado' por defecto"),
        ("5.", "Use DD-MM-YYYY o formato de fecha de Excel para fechas"),
        ("6.", "Las columnas verdes pueden dejarse vacías al importar"),
    ]
    
    for row_num, (col1, col2) in enumerate(instrucciones, 1):
        ws_instruc.cell(row=row_num, column=1, value=col1)
        ws_instruc.cell(row=row_num, column=2, value=col2)
        if "GUÍA" in col1:
            ws_instruc.cell(row=row_num, column=1).font = Font(bold=True, size=14)
    
    ws_instruc.column_dimensions['A'].width = 30
    ws_instruc.column_dimensions['B'].width = 70
    
    return wb
