"""Genera Excel de importación de OT desde registration.json de MoreApp."""
import json
import glob
import os

import openpyxl
from openpyxl.styles import Font, PatternFill

BASE = os.path.join(os.path.dirname(__file__), '..', 'Registros')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'datos_prueba')
OUT = os.path.join(OUT_DIR, 'ordenes_import_moreapp.xlsx')
OUT_DESKTOP = os.path.join(os.path.expanduser('~'), 'Desktop', 'ordenes_import_moreapp.xlsx')


def as_text(value) -> str:
    if value is None:
        return ''
    return str(value).strip()


def extraer_cliente(data: dict):
    for key in ('cliente', 'cliente1', 'CLIENTE'):
        if as_text(data.get(key)):
            return as_text(data.get(key)), '', ''

    cliente_mant = data.get('clienteParaMantenimiento') or {}
    if isinstance(cliente_mant, dict) and as_text(cliente_mant.get('NROCLIENTE')):
        return (
            as_text(cliente_mant.get('NROCLIENTE')),
            as_text(cliente_mant.get('DIRECCION')),
            as_text(cliente_mant.get('COMUNA')),
        )

    buscar = data.get('buscarCliente') or {}
    if isinstance(buscar, dict) and as_text(buscar.get('CLIENTE1')):
        return (
            as_text(buscar.get('CLIENTE1')),
            as_text(buscar.get('DIRECCION')),
            as_text(buscar.get('COMUNA')),
        )
    return '', '', ''


def extraer_tecnico(data: dict) -> str:
    for key in ('tecnicoResponsable', 'tECNICORESPONSABLE'):
        tecnico = data.get(key) or {}
        if isinstance(tecnico, dict):
            nombre = as_text(tecnico.get('NOMBRES') or tecnico.get('nombre'))
            if nombre:
                return nombre
    return ''


def resolver_tipo(data: dict, formulario: str) -> str:
    estado = as_text(data.get('estado')).upper()
    if estado == 'INCIDENCIA':
        return 'INSPECCION'

    trabajo = as_text(data.get('trabajoPrincipal'))
    actividad = as_text(data.get('actividad'))
    texto = f'{trabajo} {actividad}'.upper()

    if 'REPROGRAM' in texto or 'MANTEN' in texto or 'REPAR' in texto:
        return 'MANTENCION'
    if 'RETIRO' in texto:
        return 'RETIRO'
    if 'CAMBIO' in texto:
        return 'CAMBIO'
    if 'Registro de Medidores' in formulario or 'TELEMEDIDA' in texto or 'INSTAL' in texto:
        return 'INSTALACION'
    return 'MANTENCION'


def main():
    base = os.path.normpath(BASE)
    paths = glob.glob(os.path.join(base, '**', 'registration.json'), recursive=True)
    paths.sort(key=lambda p: (p.split(os.sep)[-3], int(p.split(os.sep)[-2]) if p.split(os.sep)[-2].isdigit() else 0))

    rows = []
    for path in paths:
        formulario = os.path.basename(os.path.dirname(os.path.dirname(path)))
        correlativo = os.path.basename(os.path.dirname(path))
        with open(path, encoding='utf-8') as handle:
            payload = json.load(handle)

        data = payload.get('data') or {}
        cliente, direccion, comuna = extraer_cliente(data)
        if not cliente:
            continue

        tecnico = extraer_tecnico(data)
        tipo = resolver_tipo(data, formulario)
        estado_json = as_text(data.get('estado')) or 'Pendiente'
        diagnostico = as_text(data.get('diagnostico'))
        trabajo = as_text(data.get('trabajoPrincipal') or data.get('actividad'))

        nombre_cliente = ''
        bloque = data.get('clienteParaMantenimiento') or data.get('buscarCliente') or {}
        if isinstance(bloque, dict):
            nombre_cliente = as_text(bloque.get('NOMBRE'))

        titulo = f'OT MoreApp #{correlativo} — {nombre_cliente or cliente}'[:200]
        descripcion = f'{formulario} | Correlativo {correlativo} | Estado terreno: {estado_json}'
        if trabajo:
            descripcion += f' | {trabajo}'
        if diagnostico:
            descripcion += f' | {diagnostico}'

        observaciones = (
            f'MoreApp submission: {as_text(payload.get("id"))} | '
            f'correlativo: {correlativo} | formulario: {formulario}'
        )

        rows.append([
            cliente,
            titulo,
            descripcion[:500],
            tipo,
            tecnico,
            'ASIGNADA',
            direccion or f'Direccion cliente {cliente}',
            comuna or 'Por definir',
            observaciones,
            correlativo,
            formulario,
        ])

    vistos_cliente = {}
    for row in rows:
        vistos_cliente.setdefault(row[0], []).append(row)
    sin_repetir = []
    for filas_cliente in vistos_cliente.values():
        filas_cliente.sort(key=lambda r: int(r[9]) if str(r[9]).isdigit() else 0, reverse=True)
        sin_repetir.append(filas_cliente[0])

    headers = [
        'Numero Cliente',
        'Titulo',
        'Descripcion',
        'Tipo Trabajo',
        'Tecnico Responsable',
        'Estado',
        'Direccion Cliente',
        'Comuna',
        'Observaciones Tecnicas',
        'Correlativo MoreApp',
        'Formulario MoreApp',
    ]

    def poblar_hoja(worksheet, filas):
        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='D9E1F2')
        for fila in filas:
            worksheet.append(fila)
        for column in worksheet.columns:
            worksheet.column_dimensions[column[0].column_letter].width = 18
        worksheet.column_dimensions['B'].width = 42
        worksheet.column_dimensions['C'].width = 55
        worksheet.column_dimensions['I'].width = 40

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    hoja_recomendada = workbook.create_sheet('Una OT por cliente', 0)
    poblar_hoja(hoja_recomendada, sin_repetir)
    hoja_completa = workbook.create_sheet('Todas las OT MoreApp', 1)
    poblar_hoja(hoja_completa, rows)

    instrucciones = workbook.create_sheet('Instrucciones', 2)
    for line in [
        ['Como importar estas OT'],
        ['1. Ir a /ordenes/ y usar Importar Excel'],
        ['2. Usa la hoja "Una OT por cliente" (primera hoja, recomendada)'],
        ['3. La hoja "Todas las OT MoreApp" tiene 88 filas; 5 clientes repetidos'],
        ['4. Si reimportas el mismo archivo, el sistema ACTUALIZA en vez de duplicar'],
        ['5. Crea clientes automaticamente si no existen'],
        [f'Filas recomendadas: {len(sin_repetir)} | Filas completas: {len(rows)}'],
    ]:
        instrucciones.append(line)

    os.makedirs(OUT_DIR, exist_ok=True)
    workbook.save(OUT)
    workbook.save(OUT_DESKTOP)
    print(f'Generado en proyecto: {OUT}')
    print(f'Copia en escritorio: {OUT_DESKTOP}')
    print(f'Filas completas: {len(rows)} | Una por cliente: {len(sin_repetir)}')
    print(f'Clientes unicos: {len(vistos_cliente)}')


if __name__ == '__main__':
    main()
