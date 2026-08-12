"""Importación masiva de cargas / órdenes de trabajo administrativas desde Excel."""
from __future__ import annotations

import math
import re
from typing import Any, Dict, List, Optional

import openpyxl
from django.db import transaction

from clientes.models import Cliente
from importaciones.models import ImportacionExcel, ImportacionExcelError
from ordenes_trabajo.models import OrdenTrabajo
from usuarios.models import Usuario

from .models import CargaAdministrativa
from .services import crear_carga

COLUMNAS = {
    'titulo': {'titulo', 'título', 'title'},
    'tipo': {'tipo', 'tipo carga', 'tipo_carga'},
    'prioridad': {'prioridad'},
    'descripcion': {'descripcion', 'descripción', 'detalle'},
    'asignado': {
        'asignado', 'asignado a', 'asignado_a', 'responsable',
        'email', 'usuario', 'rut asignado',
    },
    'cliente': {
        'cliente', 'numero cliente', 'número cliente', 'nro cliente',
        'num cliente', 'id cliente',
    },
    'orden': {'orden', 'id orden', 'ot', 'orden trabajo', 'id ot'},
    # "URL" / "Proyecto" = nombre del proyecto (listado); el ID lo genera la plataforma
    'proyecto': {
        'proyecto', 'listado proyecto', 'listado de proyecto',
        'proyecto carga', 'proyecto / carga', 'url', 'url referencia',
        'referencia', 'enlace',
    },
}

TIPOS_ALIAS = {
    'validacion ot': 'VALIDACION_OT',
    'validación ot': 'VALIDACION_OT',
    'validacion de ot': 'VALIDACION_OT',
    'validación de ot': 'VALIDACION_OT',
    'validacion_ot': 'VALIDACION_OT',
    'sci4': 'VERIFICACION_SCI4',
    'verificacion sci4': 'VERIFICACION_SCI4',
    'verificación sci4': 'VERIFICACION_SCI4',
    'actualizacion base comercial': 'VERIFICACION_SCI4',
    'actualización base comercial (sci4)': 'VERIFICACION_SCI4',
    'verificacion_sci4': 'VERIFICACION_SCI4',
    'moreapp': 'REVISION_MOREAPP',
    'revision moreapp': 'REVISION_MOREAPP',
    'revisión moreapp': 'REVISION_MOREAPP',
    'revision_moreapp': 'REVISION_MOREAPP',
    'comunicacion': 'COMUNICACION',
    'comunicación': 'COMUNICACION',
    'validacion de comunicacion': 'COMUNICACION',
    'validación de comunicación': 'COMUNICACION',
    'verificacion': 'VERIFICACION',
    'verificación': 'VERIFICACION',
    'verificacion administrativa': 'VERIFICACION',
    'verificación administrativa': 'VERIFICACION',
    'otro': 'OTRO',
    # Textos libres frecuentes del Excel administrativo → OTRO
    'actualizacion ajuste tarifario': 'OTRO',
    'actualización ajuste tarifario': 'OTRO',
    'ajuste tarifario': 'OTRO',
    'actualizacion': 'OTRO',
    'actualización': 'OTRO',
}

# Marca para verificar en el resumen de importación que corre el código nuevo
IMPORT_CARGAS_LOGIC_VERSION = 'cargas-import-v3-tipo-libre'

PRIORIDADES_ALIAS = {
    'baja': 'BAJA',
    'media': 'MEDIA',
    'alta': 'ALTA',
    'low': 'BAJA',
    'medium': 'MEDIA',
    'high': 'ALTA',
}

# Valores que en Excel se consideran “sin dato” (campos opcionales)
_VACIOS = frozenset({
    '', '-', '—', 'n/a', 'na', 'n.a.', 'n.a', 'null', 'none', 'nil',
    'sin asignar', 's/a', 's/n', 'sn', 'no aplica', 'ninguno', 'ninguna',
})


def _es_vacio(valor: str) -> bool:
    return _limpiar_header(valor) in _VACIOS


def _limpiar_header(valor) -> str:
    if valor is None:
        return ''
    texto = str(valor).strip().lower()
    texto = re.sub(r'\s+', ' ', texto)
    return texto


def _mapear_headers(headers: List[Any]) -> Dict[str, int]:
    indice: Dict[str, int] = {}
    for i, raw in enumerate(headers):
        clave = _limpiar_header(raw)
        if not clave:
            continue
        for campo, alias in COLUMNAS.items():
            if clave in alias and campo not in indice:
                indice[campo] = i
                break
    return indice


def _valor_fila(valores: List[Any], indice: Dict[str, int], campo: str) -> str:
    if campo not in indice:
        return ''
    i = indice[campo]
    if i >= len(valores):
        return ''
    v = valores[i]
    if v is None:
        return ''
    if isinstance(v, float):
        if math.isnan(v):
            return ''
        if v.is_integer():
            return str(int(v))
    texto = str(v).strip()
    if _es_vacio(texto):
        return ''
    return texto


def _fila_a_texto(headers: List[Any], valores: List[Any]) -> str:
    partes = []
    for h, v in zip(headers, valores):
        if h is None and (v is None or str(v).strip() == ''):
            continue
        partes.append(f'{h}={v}')
    return ' | '.join(partes)[:2000]


def _resolver_tipo(raw: str):
    """
    Resuelve el tipo de carga.
    Retorna (codigo, texto_libre_opcional).
    NUNCA lanza excepción: textos libres (p. ej. «Actualización Ajuste Tarifario»)
    se guardan como OTRO y el valor original va a la descripción.
    """
    try:
        if not raw or _es_vacio(raw):
            return 'VERIFICACION', None

        texto = str(raw).strip()
        codigo = texto.upper().replace(' ', '_')
        validos = {c[0] for c in CargaAdministrativa.TIPO_CHOICES}
        if codigo in validos:
            return codigo, None

        clave = _limpiar_header(texto)
        alias = TIPOS_ALIAS.get(clave)
        if alias:
            # Si el alias es OTRO por texto libre conocido, conservar el texto original
            if alias == 'OTRO' and clave not in {'otro'}:
                return 'OTRO', texto
            return alias, None

        for code, label in CargaAdministrativa.TIPO_CHOICES:
            if _limpiar_header(label) == clave:
                return code, None

        return 'OTRO', texto
    except Exception:
        # Último recurso: nunca bloquear la fila por el tipo
        fallback = str(raw).strip() if raw is not None else ''
        return 'OTRO', fallback or None


def _resolver_prioridad(raw: str) -> str:
    if not raw or _es_vacio(raw):
        return 'MEDIA'
    codigo = raw.strip().upper()
    validos = {p[0] for p in CargaAdministrativa.PRIORIDAD_CHOICES}
    if codigo in validos:
        return codigo
    alias = PRIORIDADES_ALIAS.get(_limpiar_header(raw))
    if alias:
        return alias
    # Prioridad desconocida: no falla, usa MEDIA
    return 'MEDIA'


def _texto_asignado(raw: str) -> str:
    """Asignado es texto libre del Excel (opcional). No exige usuario del sistema."""
    if not raw or _es_vacio(raw):
        return ''
    return raw.strip()[:255]


def _intentar_usuario_asignado(texto: str) -> Optional[Usuario]:
    """Si el texto coincide con un ADMIN/ADMINISTRATIVO, enlaza el usuario (opcional)."""
    if not texto:
        return None
    qs = Usuario.objects.filter(rol__in=['ADMIN', 'ADMINISTRATIVO'], is_active=True)
    return (
        qs.filter(email__iexact=texto).first()
        or qs.filter(nombre_interno__iexact=texto).first()
        or qs.filter(rut__iexact=texto).first()
        or qs.filter(nombre__iexact=texto).first()
    )


def _resolver_cliente(raw: str) -> Optional[Cliente]:
    """Opcional: vacío → sin cliente. Solo valida si viene un valor."""
    if not raw or _es_vacio(raw):
        return None
    texto = raw.strip()
    cliente = Cliente.objects.filter(numero_cliente__iexact=texto, activo=True).first()
    if not cliente and texto.isdigit():
        cliente = Cliente.objects.filter(pk=int(texto), activo=True).first()
    if not cliente:
        raise ValueError(
            f'Cliente «{texto}» no encontrado o inactivo. '
            'Si no aplica, deja la celda vacía.'
        )
    return cliente


def _resolver_orden(raw: str) -> Optional[OrdenTrabajo]:
    """Opcional: vacío → sin OT. Solo valida si viene un valor."""
    if not raw or _es_vacio(raw):
        return None
    texto = raw.strip()
    if not texto.isdigit():
        raise ValueError(
            f'ID Orden inválido: «{texto}». Debe ser numérico, '
            'o deja la celda vacía si no aplica.'
        )
    orden = OrdenTrabajo.objects.filter(pk=int(texto), eliminado=False).first()
    if not orden:
        raise ValueError(
            f'Orden de trabajo #{texto} no encontrada. '
            'Si no aplica, deja la celda vacía.'
        )
    return orden


def _clave_duplicado(titulo: str) -> str:
    """La identificación de duplicados es solo por título (único distintivo)."""
    return (titulo or '').strip().casefold()


def _buscar_duplicado_db(titulo: str) -> Optional[CargaAdministrativa]:
    return (
        CargaAdministrativa.objects.filter(
            eliminado=False,
            titulo__iexact=(titulo or '').strip(),
        )
        .order_by('-fecha_creacion')
        .first()
    )


def importar_cargas_excel(archivo, usuario) -> ImportacionExcel:
    """
    Crea cargas administrativas desde Excel.
    El ID (# correlativo) lo genera automáticamente la plataforma.
    Duplicados se detectan solo por Título (no se sobrescriben).
    La columna Proyecto/URL guarda el listado de proyecto asociado.
    """
    importacion = ImportacionExcel.objects.create(
        tipo='CARGAS_ADMINISTRATIVAS',
        archivo_original=getattr(archivo, 'name', 'Upload'),
        usuario=usuario,
    )

    try:
        nombre_archivo = getattr(archivo, 'name', '') or ''
        if nombre_archivo.lower().endswith('.xls') and not nombre_archivo.lower().endswith('.xlsx'):
            raise ValueError(
                'Formato .xls no soportado. Guarda el archivo como Excel (.xlsx) e intenta de nuevo.'
            )

        if hasattr(archivo, 'seek'):
            archivo.seek(0)
        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        indice = _mapear_headers(headers)

        if 'titulo' not in indice:
            raise ValueError(
                'El Excel debe incluir la columna «Titulo». '
                f'Columnas detectadas: {", ".join(str(h) for h in headers if h)}'
            )

        contador = 0
        exitosas = 0
        errores = 0
        duplicados = 0
        claves_archivo: set = set()

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            valores = [cell.value for cell in row]
            if not any(v is not None and str(v).strip() != '' for v in valores):
                continue

            titulo_raw = _valor_fila(valores, indice, 'titulo')
            if titulo_raw and _limpiar_header(titulo_raw) in COLUMNAS['titulo']:
                # Fila de encabezado repetida
                continue

            contador += 1
            try:
                titulo = titulo_raw
                if not titulo:
                    raise ValueError('Titulo es obligatorio')

                tipo_raw = _valor_fila(valores, indice, 'tipo')
                try:
                    tipo, tipo_libre = _resolver_tipo(tipo_raw)
                except Exception:
                    tipo, tipo_libre = 'OTRO', (tipo_raw or None)
                # Compatibilidad si por algún motivo quedó una versión vieja que devolvía solo str
                if not isinstance(tipo, str):
                    tipo, tipo_libre = 'OTRO', tipo_raw or None
                prioridad = _resolver_prioridad(_valor_fila(valores, indice, 'prioridad'))
                descripcion = _valor_fila(valores, indice, 'descripcion')
                if tipo_libre:
                    nota_tipo = f'Tipo (Excel): {tipo_libre}'
                    descripcion = f'{nota_tipo}\n{descripcion}' if descripcion else nota_tipo
                asignado_texto = _texto_asignado(_valor_fila(valores, indice, 'asignado'))
                asignado_usuario = _intentar_usuario_asignado(asignado_texto)
                cliente = _resolver_cliente(_valor_fila(valores, indice, 'cliente'))
                orden = _resolver_orden(_valor_fila(valores, indice, 'orden'))
                proyecto = _valor_fila(valores, indice, 'proyecto')

                clave = _clave_duplicado(titulo)
                if clave in claves_archivo:
                    duplicados += 1
                    ImportacionExcelError.objects.create(
                        importacion=importacion,
                        numero_fila=idx,
                        motivo=(
                            'Duplicado en el archivo: ya hay otra fila con el mismo Título. '
                            'El título identifica de forma única cada carga administrativa.'
                        ),
                        data_cruda=_fila_a_texto(headers, valores),
                    )
                    continue

                existente = _buscar_duplicado_db(titulo)
                if existente:
                    duplicados += 1
                    ImportacionExcelError.objects.create(
                        importacion=importacion,
                        numero_fila=idx,
                        motivo=(
                            f'Duplicado: ya existe la carga #{existente.pk} '
                            f'con el mismo título «{existente.titulo}» '
                            f'(estado {existente.get_estado_display()}). '
                            'No se sobrescribe. El ID lo asigna la plataforma.'
                        ),
                        data_cruda=_fila_a_texto(headers, valores),
                    )
                    continue

                with transaction.atomic():
                    crear_carga(
                        usuario,
                        titulo=titulo,
                        tipo=tipo,
                        descripcion=descripcion,
                        prioridad=prioridad,
                        asignado_a=asignado_usuario,
                        asignado_texto=asignado_texto,
                        orden=orden,
                        cliente=cliente,
                        proyecto=proyecto,
                    )
                claves_archivo.add(clave)
                exitosas += 1
            except Exception as exc:
                errores += 1
                ImportacionExcelError.objects.create(
                    importacion=importacion,
                    numero_fila=idx,
                    motivo=str(exc),
                    data_cruda=_fila_a_texto(headers, valores),
                )

        importacion.total_filas = contador
        importacion.exitosas = exitosas
        importacion.fallidas = errores + duplicados
        importacion.estado = (
            'COMPLETADO' if contador > 0 and exitosas > 0
            else ('ERROR' if contador == 0 else 'COMPLETADO')
        )
        if contador == 0:
            importacion.observaciones = (
                f'No se encontraron filas con datos para importar. '
                f'[{IMPORT_CARGAS_LOGIC_VERSION}]'
            )
        else:
            importacion.observaciones = (
                f'Total de registros encontrados: {contador}. '
                f'Registros cargados correctamente: {exitosas}. '
                f'Registros con errores: {errores}. '
                f'Registros duplicados: {duplicados}. '
                f'[{IMPORT_CARGAS_LOGIC_VERSION}]'
            )
        # Metadatos extra para la respuesta JSON (parseables)
        importacion.observaciones += f'\n[meta] errores={errores};duplicados={duplicados}'
        importacion.save()
        # Adjuntar conteos en el objeto en memoria para la vista
        importacion._conteo_errores = errores  # type: ignore[attr-defined]
        importacion._conteo_duplicados = duplicados  # type: ignore[attr-defined]
    except Exception as exc:
        importacion.estado = 'ERROR'
        importacion.observaciones = f'Error en importación: {exc}'
        importacion.save()
        importacion._conteo_errores = 0  # type: ignore[attr-defined]
        importacion._conteo_duplicados = 0  # type: ignore[attr-defined]

    return importacion


def resumen_importacion(importacion: ImportacionExcel) -> Dict[str, int]:
    """Extrae conteos de errores/duplicados desde observaciones o atributos."""
    errores = getattr(importacion, '_conteo_errores', None)
    duplicados = getattr(importacion, '_conteo_duplicados', None)
    if errores is not None and duplicados is not None:
        return {'errores': int(errores), 'duplicados': int(duplicados)}

    obs = importacion.observaciones or ''
    m_err = re.search(r'Registros con errores:\s*(\d+)', obs)
    m_dup = re.search(r'Registros duplicados:\s*(\d+)', obs)
    if m_err and m_dup:
        return {'errores': int(m_err.group(1)), 'duplicados': int(m_dup.group(1))}

    # Fallback: clasificar errores guardados
    dups = 0
    errs = 0
    for e in importacion.errores.all():
        if (e.motivo or '').startswith('Duplicado'):
            dups += 1
        else:
            errs += 1
    return {'errores': errs, 'duplicados': dups}


def exportar_cargas_excel(cargas):
    """Genera workbook Excel con las cargas administrativas recibidas."""
    from importaciones.utils import aplicar_estilo_hoja_exportacion
    from ordenes_trabajo.observaciones_html import observaciones_a_texto_plano

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cargas administrativas'
    # ID Carga es solo informativo (correlativo de la plataforma); no se usa al reimportar.
    ws.append([
        'ID Carga',
        'Titulo',
        'Tipo',
        'Prioridad',
        'Estado',
        'Descripcion',
        'Asignado',
        'Cliente',
        'ID Orden',
        'Proyecto',
        'Observaciones',
        'Creado por',
        'Fecha creacion',
        'Fecha asignacion',
        'Fecha completada',
    ])

    for carga in cargas:
        ws.append([
            carga.id,
            carga.titulo or '',
            carga.get_tipo_display(),
            carga.get_prioridad_display(),
            carga.get_estado_display(),
            carga.descripcion or '',
            carga.asignado_a.nombre_interno if carga.asignado_a_id else (carga.asignado_texto or ''),
            carga.cliente.numero_cliente if carga.cliente_id else '',
            carga.orden_id or '',
            carga.proyecto or '',
            observaciones_a_texto_plano(carga.observaciones or ''),
            carga.creado_por.nombre_interno if carga.creado_por_id else '',
            carga.fecha_creacion.strftime('%d/%m/%Y %H:%M') if carga.fecha_creacion else '',
            carga.fecha_asignacion.strftime('%d/%m/%Y %H:%M') if carga.fecha_asignacion else '',
            carga.fecha_completada.strftime('%d/%m/%Y %H:%M') if carga.fecha_completada else '',
        ])

    # Filtro en Tipo / Prioridad / Estado / Asignado / Cliente / Proyecto (cols 3–10).
    aplicar_estilo_hoja_exportacion(ws, auto_filter=True, filter_from_col=3, filter_to_col=10)
    return wb
