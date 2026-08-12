from io import BytesIO

import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import Client, TestCase
from django.urls import reverse

from clientes.models import Cliente
from importaciones.models import ImportacionExcel
from usuarios.models import Usuario

from .import_excel import importar_cargas_excel, resumen_importacion
from .models import AdjuntoCarga, CargaAdministrativa
from .services import crear_carga, eliminar_carga


def _xlsx_bytes(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class CargasImportDeleteTests(TestCase):
    def setUp(self):
        self.password = 'admin1234'
        self.admin = Usuario.objects.create_user(
            rut='15151515-1',
            email='admin_imp@delco.cl',
            password=self.password,
            nombre='Admin',
            apellido='Imp',
            nombre_interno='admin_imp',
            rol='ADMIN',
            is_active=True,
            is_staff=True,
        )
        self.admin_op = Usuario.objects.create_user(
            rut='16161616-1',
            email='admvo_imp@delco.cl',
            password=self.password,
            nombre='Admvo',
            apellido='Imp',
            nombre_interno='admvo_imp',
            rol='ADMINISTRATIVO',
            is_active=True,
        )
        self.tecnico = Usuario.objects.create_user(
            rut='17171717-1',
            email='tec_imp@delco.cl',
            password=self.password,
            nombre='Tec',
            apellido='Imp',
            nombre_interno='tec_imp',
            rol='TECNICO',
            is_active=True,
        )
        self.cliente = Cliente.objects.create(
            numero_cliente='CLI-IMP-001',
            direccion='Dir',
            comuna='Santiago',
            tipo_suministro='ELECTRICO',
            sector='CENTRO',
            customer_name='Cliente Imp',
            installation_address='Inst',
            meter_manufacturer_id='TEST',
            meter_serial_n_1='SER-IMP-001',
            activo=True,
        )
        self.client = Client()

    def test_importar_excel_crea_y_reporta_errores_duplicados(self):
        crear_carga(
            self.admin,
            titulo='Ya existe',
            tipo='VERIFICACION',
            cliente=self.cliente,
        )
        buf = _xlsx_bytes([
            ['Titulo', 'Tipo', 'Prioridad', 'Cliente', 'Asignado', 'Descripcion', 'Proyecto'],
            ['Nueva carga A', 'VERIFICACION', 'ALTA', 'CLI-IMP-001', 'admvo_imp', 'ok', 'Proyecto Norte'],
            ['', 'VERIFICACION', 'MEDIA', 'CLI-IMP-001', '', 'sin titulo', ''],
            ['Ya existe', 'VERIFICACION', 'MEDIA', 'CLI-IMP-001', '', 'dup db', ''],
            ['Nueva carga A', 'VERIFICACION', 'ALTA', 'CLI-IMP-001', '', 'dup file', ''],
            ['Otra ok', 'OTRO', 'BAJA', '', 'admvo_imp@delco.cl', 'segunda', 'Proyecto Sur'],
            ['Con ID ignorado', 'OTRO', 'MEDIA', '', '', 'id no importa', 'Alpha'],
        ])
        # Columna ID Carga presente: debe ignorarse (no error ni bloqueo)
        buf2 = _xlsx_bytes([
            ['ID Carga', 'Titulo', 'URL'],
            ['99999', 'Desde URL como proyecto', 'Listado Beta'],
        ])
        archivo = SimpleUploadedFile(
            'cargas.xlsx',
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        importacion = importar_cargas_excel(archivo, self.admin)
        conteos = resumen_importacion(importacion)

        self.assertEqual(importacion.total_filas, 6)
        self.assertEqual(importacion.exitosas, 3)
        self.assertEqual(conteos['errores'], 1)
        self.assertEqual(conteos['duplicados'], 2)
        self.assertIn('Registros cargados correctamente: 3', importacion.observaciones)
        carga_a = CargaAdministrativa.objects.get(
            eliminado=False, titulo='Nueva carga A', tipo='VERIFICACION'
        )
        self.assertEqual(carga_a.proyecto, 'Proyecto Norte')
        self.assertIn('proyecto=', carga_a.url_referencia)
        self.assertTrue(
            CargaAdministrativa.objects.filter(eliminado=False, titulo='Otra ok').exists()
        )
        # No se creó una segunda "Ya existe"
        self.assertEqual(
            CargaAdministrativa.objects.filter(
                eliminado=False, titulo__iexact='Ya existe'
            ).count(),
            1,
        )

        archivo_id = SimpleUploadedFile(
            'cargas_id.xlsx',
            buf2.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        imp2 = importar_cargas_excel(archivo_id, self.admin)
        self.assertEqual(imp2.exitosas, 1)
        carga_url = CargaAdministrativa.objects.get(titulo='Desde URL como proyecto')
        self.assertEqual(carga_url.proyecto, 'Listado Beta')
        self.assertNotEqual(carga_url.pk, 99999)

    def test_importar_solo_titulo_sin_asignado_cliente_orden(self):
        """Asignado, Cliente e ID Orden son opcionales (pueden omitirse o ir vacíos)."""
        buf = _xlsx_bytes([
            ['Titulo', 'Proyecto', 'Asignado', 'Cliente', 'ID Orden'],
            ['Solo titulo A', 'Proyecto X', '', '', ''],
            ['Solo titulo B', 'Proyecto Y', '-', 'n/a', 'N/A'],
            ['Solo titulo C', '', '', '', ''],
        ])
        # Sin columnas opcionales en absoluto
        buf_min = _xlsx_bytes([
            ['Titulo'],
            ['Minimo uno'],
            ['Minimo dos'],
        ])
        archivo = SimpleUploadedFile(
            'solo_titulo.xlsx',
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        imp = importar_cargas_excel(archivo, self.admin)
        self.assertEqual(imp.exitosas, 3)
        self.assertEqual(imp.fallidas, 0)
        a = CargaAdministrativa.objects.get(titulo='Solo titulo A')
        self.assertIsNone(a.asignado_a_id)
        self.assertIsNone(a.cliente_id)
        self.assertIsNone(a.orden_id)
        self.assertEqual(a.proyecto, 'Proyecto X')

        archivo_min = SimpleUploadedFile(
            'min.xlsx',
            buf_min.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        imp2 = importar_cargas_excel(archivo_min, self.admin)
        self.assertEqual(imp2.exitosas, 2)
        self.assertEqual(imp2.fallidas, 0)

    def test_importar_tipo_texto_libre_no_falla(self):
        """Tipos del Excel que no son códigos del sistema se guardan como OTRO."""
        buf = _xlsx_bytes([
            ['Titulo', 'Tipo', 'Descripcion'],
            ['Ajuste 1', 'Actualización Ajuste Tarifario', 'Detalle del ajuste'],
            ['Ajuste 2', 'Actualización Ajuste Tarifario', ''],
            ['Validacion formal', 'Validación de OT', 'usa etiqueta'],
        ])
        archivo = SimpleUploadedFile(
            'tipos_libres.xlsx',
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        imp = importar_cargas_excel(archivo, self.admin)
        self.assertEqual(imp.exitosas, 3)
        self.assertEqual(imp.fallidas, 0)

        c1 = CargaAdministrativa.objects.get(titulo='Ajuste 1')
        self.assertEqual(c1.tipo, 'OTRO')
        self.assertIn('Actualización Ajuste Tarifario', c1.descripcion)
        self.assertIn('Detalle del ajuste', c1.descripcion)

        c2 = CargaAdministrativa.objects.get(titulo='Ajuste 2')
        self.assertEqual(c2.tipo, 'OTRO')
        self.assertIn('Actualización Ajuste Tarifario', c2.descripcion)

        c3 = CargaAdministrativa.objects.get(titulo='Validacion formal')
        self.assertEqual(c3.tipo, 'VALIDACION_OT')

    def test_importar_via_vista_json(self):
        self.assertTrue(self.client.login(rut=self.admin_op.rut, password=self.password))
        buf = _xlsx_bytes([
            ['Titulo', 'Tipo'],
            ['Desde vista', 'COMUNICACION'],
        ])
        response = self.client.post(
            reverse('cargas_importar'),
            {
                'archivo': SimpleUploadedFile(
                    'c.xlsx',
                    buf.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
            },
            HTTP_X_REQUESTED_WITH='XMLHttpRequest',
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data['success'])
        self.assertEqual(data['exitosas'], 1)
        self.assertEqual(data['duplicados'], 0)
        self.assertTrue(ImportacionExcel.objects.filter(tipo='CARGAS_ADMINISTRATIVAS').exists())

    def test_tecnico_no_puede_importar_ni_eliminar(self):
        carga = crear_carga(self.admin, titulo='Protegida', tipo='OTRO')
        self.assertTrue(self.client.login(rut=self.tecnico.rut, password=self.password))
        r1 = self.client.post(reverse('cargas_importar'), {})
        self.assertEqual(r1.status_code, 403)
        r2 = self.client.post(reverse('cargas_eliminar', kwargs={'pk': carga.pk}))
        self.assertEqual(r2.status_code, 403)
        carga.refresh_from_db()
        self.assertFalse(carga.eliminado)

    def test_eliminar_individual_conserva_adjuntos(self):
        carga = crear_carga(self.admin, titulo='Con evidencia', tipo='OTRO')
        adj = AdjuntoCarga.objects.create(
            carga=carga,
            tipo='FOTO',
            nombre_archivo='ev.png',
            subido_por=self.admin,
        )
        self.assertTrue(self.client.login(rut=self.admin_op.rut, password=self.password))
        response = self.client.post(reverse('cargas_eliminar', kwargs={'pk': carga.pk}))
        self.assertEqual(response.status_code, 302)
        carga.refresh_from_db()
        self.assertTrue(carga.eliminado)
        self.assertEqual(carga.eliminado_por_id, self.admin_op.pk)
        adj.refresh_from_db()
        self.assertFalse(adj.eliminado)
        # No aparece en listado ni detalle
        self.assertTrue(self.client.login(rut=self.admin.rut, password=self.password))
        lista = self.client.get(reverse('cargas_list'))
        self.assertEqual(lista.status_code, 200)
        ids_listado = [c.pk for c in lista.context['cargas']]
        self.assertNotIn(carga.pk, ids_listado)
        detalle = self.client.get(reverse('cargas_detalle', kwargs={'pk': carga.pk}))
        self.assertEqual(detalle.status_code, 404)

    def test_eliminar_masivo(self):
        c1 = crear_carga(self.admin, titulo='Masiva 1', tipo='OTRO')
        c2 = crear_carga(self.admin, titulo='Masiva 2', tipo='OTRO')
        self.assertTrue(self.client.login(rut=self.admin.rut, password=self.password))
        response = self.client.post(
            reverse('cargas_eliminar_masivo'),
            {'ids': [str(c1.pk), str(c2.pk)]},
        )
        self.assertEqual(response.status_code, 302)
        c1.refresh_from_db()
        c2.refresh_from_db()
        self.assertTrue(c1.eliminado)
        self.assertTrue(c2.eliminado)

    def test_soft_delete_service_idempotente(self):
        carga = crear_carga(self.admin, titulo='Idem', tipo='OTRO')
        self.assertTrue(eliminar_carga(carga, self.admin, motivo='prueba'))
        self.assertFalse(eliminar_carga(carga, self.admin, motivo='otra'))

    def test_exportar_excel_con_filtros(self):
        from .import_excel import exportar_cargas_excel

        c1 = crear_carga(self.admin, titulo='Export Alta', tipo='VERIFICACION', prioridad='ALTA')
        crear_carga(self.admin, titulo='Export Baja', tipo='OTRO', prioridad='BAJA')
        crear_carga(self.admin, titulo='Export Sci4', tipo='VERIFICACION_SCI4')

        self.assertTrue(self.client.login(rut=self.admin_op.rut, password=self.password))

        # Filtrado por tipo
        response = self.client.get(reverse('cargas_exportar'), {'filtrar': '1', 'tipo': 'VERIFICACION'})
        self.assertEqual(response.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            response['Content-Type'],
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        titulos = [row[1].value for row in ws.iter_rows(min_row=2, values_only=False)]
        self.assertIn('Export Alta', titulos)
        self.assertNotIn('Export Baja', titulos)
        self.assertNotIn('Export Sci4', titulos)
        self.assertIsNotNone(ws.auto_filter.ref)

        # Todas
        response_todas = self.client.get(reverse('cargas_exportar'), {'todas': '1'})
        self.assertEqual(response_todas.status_code, 200)
        wb2 = openpyxl.load_workbook(BytesIO(response_todas.content))
        titulos2 = [row[1].value for row in wb2.active.iter_rows(min_row=2, values_only=False)]
        self.assertIn('Export Alta', titulos2)
        self.assertIn('Export Baja', titulos2)
        self.assertIn('Export Sci4', titulos2)

        # Helper directo incluye ID
        wb3 = exportar_cargas_excel([c1])
        self.assertEqual(wb3.active.cell(2, 1).value, c1.pk)

    def test_tecnico_no_puede_exportar(self):
        self.assertTrue(self.client.login(rut=self.tecnico.rut, password=self.password))
        response = self.client.get(reverse('cargas_exportar'), {'todas': '1'})
        self.assertEqual(response.status_code, 403)
