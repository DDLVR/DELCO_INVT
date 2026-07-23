"""Generación de Excel y PDF para reportes del PDF punto 9."""

from __future__ import annotations

from io import BytesIO
from typing import Iterable, List, Optional, Sequence

import openpyxl
from django.http import HttpResponse
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from importaciones.utils import aplicar_estilo_hoja_exportacion
from web.services.export_filenames import nombre_exportacion_con_fecha


class PdfExportUnavailable(Exception):
    """Se lanza cuando falta reportlab u otra dependencia de PDF."""


def _as_cell_text(value) -> str:
    if value is None:
        return ''
    return str(value)


def build_excel_response(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    title: Optional[str] = None,
    sheet_title: str = 'Reporte',
    group_by_first_column: bool = False,
) -> HttpResponse:
    """
    Excel de reportes con el mismo look & feel que exportaciones de clientes/OT:
    título Delco, encabezado azul, zebra, bordes y auto-filtro.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_title or 'Reporte')[:31]

    row_list = [list(r) for r in rows]
    generado = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')
    titulo = (title or 'Reporte operativo').strip() or 'Reporte operativo'

    # Bloque de portada (misma idea visual que el PDF de reportes)
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=14, color='1F4E79', name='Calibri')
    ws['A1'].alignment = Alignment(vertical='center')
    ws['A2'] = f'DelcoChile · Generado {generado} · {len(row_list)} registro(s)'
    ws['A2'].font = Font(size=10, color='666666', name='Calibri')
    ws['A2'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=header)

    for offset, row in enumerate(row_list):
        excel_row = header_row + 1 + offset
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=excel_row, column=col_idx, value=value if value is not None else '')

    # Merge título/meta al ancho de columnas
    last_col = max(len(headers), 1)
    last_letter = get_column_letter(last_col)
    ws.merge_cells(f'A1:{last_letter}1')
    ws.merge_cells(f'A2:{last_letter}2')

    aplicar_estilo_hoja_exportacion(
        ws,
        header_row=header_row,
        freeze=f'A{header_row + 1}',
        auto_filter=True,
        filter_from_col=1,
        filter_to_col=min(4, last_col),
        max_width=42,
    )

    # Resaltar grupos del primer valor (p. ej. misma IP / mismo medidor)
    if group_by_first_column and row_list and last_col >= 1:
        fills = (
            PatternFill(start_color='FFF4E5', end_color='FFF4E5', fill_type='solid'),
            PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'),
        )
        grupo = -1
        valor_prev = object()
        for offset, row in enumerate(row_list):
            valor = row[0] if row else None
            if valor != valor_prev:
                grupo += 1
                valor_prev = valor
            fill = fills[grupo % 2]
            excel_row = header_row + 1 + offset
            # Solo columna clave + contador (si existe)
            for col_idx in range(1, min(3, last_col) + 1):
                cell = ws.cell(row=excel_row, column=col_idx)
                # No pisa el zebra completo: marca la clave del problema
                cell.fill = fill
                if col_idx == 1:
                    cell.font = Font(bold=True, size=10, name='Calibri', color='212121')

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = nombre_exportacion_con_fecha(filename)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def build_pdf_response(
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence],
    title: Optional[str] = None,
) -> HttpResponse:
    """PDF tabular en landscape A4; celdas largas se truncan para legibilidad."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:
        raise PdfExportUnavailable(
            'No está instalado el paquete reportlab en el servidor. '
            'Ejecuta: pip install "reportlab>=3.6.0,<4.0"'
        ) from exc

    buffer = BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title or filename,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading2'],
        fontSize=12,
        alignment=TA_CENTER,
        spaceAfter=4,
    )
    meta_style = ParagraphStyle(
        'ReportMeta',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_CENTER,
        textColor=colors.grey,
        spaceAfter=8,
    )
    cell_style = ParagraphStyle(
        'ReportCell',
        parent=styles['Normal'],
        fontSize=7,
        leading=9,
        alignment=TA_LEFT,
        wordWrap='CJK',
    )
    header_style = ParagraphStyle(
        'ReportHeader',
        parent=cell_style,
        fontName='Helvetica-Bold',
        fontSize=7,
        leading=9,
    )

    max_chars = 48 if len(headers) <= 8 else (36 if len(headers) <= 12 else 28)

    def _cell_paragraph(text: str, style) -> Paragraph:
        raw = _as_cell_text(text).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        if len(raw) > max_chars:
            raw = raw[: max_chars - 1] + '…'
        return Paragraph(raw or '—', style)

    header_row = [_cell_paragraph(h, header_style) for h in headers]
    data: List[List] = [header_row]
    row_list = list(rows)
    for row in row_list:
        cells = list(row) + [''] * max(0, len(headers) - len(row))
        data.append([
            _cell_paragraph(cells[i] if i < len(cells) else '', cell_style)
            for i in range(len(headers))
        ])

    usable_width = page_size[0] - doc.leftMargin - doc.rightMargin
    col_width = usable_width / max(len(headers), 1)
    table = Table(data, colWidths=[col_width] * len(headers), repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f6fa')]),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#c5d0de')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]
        )
    )

    generado = timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')
    story = [
        Paragraph(_as_cell_text(title or 'Reporte operativo'), title_style),
        Paragraph(f'DelcoChile · Generado {generado} · {len(row_list)} registro(s)', meta_style),
        Spacer(1, 4),
        table,
    ]
    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    filename = nombre_exportacion_con_fecha(filename)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
